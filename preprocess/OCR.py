import warnings
warnings.filterwarnings('ignore')
import unstructured_pytesseract

unstructured_pytesseract.pytesseract.tesseract_cmd = r'C:\Users\Divyesh Parmar\Downloads\Tesserect OCR\tesseract.exe'

import pandas as pd
import os

from unstructured.partition.pdf import partition_pdf
import unstructured.partition

help(unstructured.partition)

from openpyxl import Workbook, load_workbook

# Specify the path to your PDF file
# filename = "scanned_rotated_right_enhanced.pdf"
filename = "scanned4.pdf"

# Call the partition_pdf function
# Returns a List[Element] present in the pages of the parsed pdf document
elements = partition_pdf(filename=filename,
                         infer_table_structure=True,
                          extract_images_in_pdf=True, 
                            languages=["eng"],
                           extract_image_block_types=["Table"],
                         strategy='hi_res',
           )

tables = [el for el in elements if el.category == "Table"]

dfs = []

excel_files = []
for i, table in enumerate(tables):
    # Get HTML content of the table
    table_html = table.metadata.text_as_html
    # Parse HTML to DataFrame
    parsed_dfs = pd.read_html(table_html)

    #Code for exporting into different excel and then combining them
    if parsed_dfs:
        df = parsed_dfs[0]
        
        df = df.replace(r'[\|\[\]]', '', regex=True)
        # df = df.applymap(
        #     lambda x: str(x).replace("|", "")
        #     .replace("[", "")
        #     .replace("]", "")
        #     if isinstance(x, str) else x
        #     )
        print(f"{i} Table in Dataframes : \n", df)
        # dfs.append(df)
        output_filename = f"output_table_{i+1}.xlsx"
        df.to_excel(output_filename, index=True)
        excel_files.append(output_filename)

    else:
        print(f"No table parsed from element {i}")

dataframes = [pd.read_excel(f, engine='openpyxl') for f in excel_files]


#Merge all dataframe into one excel sheet
wb = Workbook()
ws = wb.active
current_row = 1

for df in dataframes:
    for col_num, column_title in enumerate(df.columns, 1):
        ws.cell(row=current_row, column=col_num, value=column_title)

    for row_num, row_data in enumerate(df.itertuples(index=False), start=2):
        current_row += 1
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=current_row, column=col_num, value=value)
    current_row += 2

wb.save(f"{filename.replace('.pdf', '')}_tables.xlsx")


#Deleting temporary files
for f in excel_files:
    try:
        os.remove(f)
        print(f"Deleted temp file: {f}")
    except Exception as e:
        print(f"Error deleting {f}: {e}")
