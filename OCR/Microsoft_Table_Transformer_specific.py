from __future__ import annotations

import os
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from threading import Lock
from typing import Callable

# Keep each Tesseract worker single-threaded so parallel cell OCR can use the CPU.
os.environ.setdefault("OMP_THREAD_LIMIT", "1")

import fitz  # PyMuPDF
import numpy as np
import pytesseract
import torch
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageOps
from transformers import AutoImageProcessor, TableTransformerForObjectDetection

from env_config import get_input_pdf


# =============================================================================
# CONFIGURATION
# =============================================================================

INPUT_PDF = get_input_pdf()
OUTPUT_XLSX = Path("extracted_tables.xlsx")
DEBUG_DIR = Path(__file__).resolve().parent / "table_debug"

DETECTION_MODEL_NAME = "microsoft/table-transformer-detection"
STRUCTURE_MODEL_NAME = (
    "microsoft/table-transformer-structure-recognition-v1.1-all"
)

PDF_DPI = 250
DETECTION_THRESHOLD = 0.75
STRUCTURE_THRESHOLD = 0.60

TABLE_PADDING = 12
CELL_INSET = 2

# Parallel OCR workers. Accuracy-neutral; only changes throughput.
OCR_WORKERS = max(1, min(8, (os.cpu_count() or 4)))
# Stop trying more OCR variants once a strong result is found.
OCR_EARLY_EXIT_CONFIDENCE = 70.0

# Choose text OCR engine: "tesseract" or "easyocr"
OCR_ENGINE = "easyocr"
# Process multiple PDF pages concurrently. Keep this modest because each page
# worker holds its own table-detection/structure models in memory.
PAGE_WORKERS = max(1, min(2, (os.cpu_count() or 4)))

# Uncomment and update this on Windows only when tesseract.exe is not in PATH.
# This is unnecessary after replacing default_ocr() with your existing OCR.
#
pytesseract.pytesseract.tesseract_cmd = (
    r"C:\Users\Divyesh Parmar\Downloads\Tesserect OCR\tesseract.exe"
)


def effective_ocr_workers() -> int:
    """
    Return the useful OCR parallelism for the selected engine.

    EasyOCR on CPU is internally heavy and uses a shared reader lock below, so
    high outer parallelism only adds contention and startup overhead.
    """
    engine = str(OCR_ENGINE or "tesseract").strip().lower()
    if engine == "easyocr":
        return min(2, OCR_WORKERS) if torch.cuda.is_available() else 1
    return OCR_WORKERS


def effective_page_workers() -> int:
    """
    Return page-level parallelism.

    GPU model inference should stay single-page to avoid contention/VRAM spikes.
    CPU inference can benefit from a small number of concurrent page workers.
    """
    return 1 if torch.cuda.is_available() else PAGE_WORKERS


@dataclass(frozen=True)
class Detection:
    label: str
    score: float
    box: tuple[float, float, float, float]


# =============================================================================
# GENERAL GEOMETRY
# =============================================================================

def safe_sheet_name(name: str) -> str:
    """Return a legal Excel worksheet name."""
    name = re.sub(r"[:\\/?*\[\]]", "_", name)
    return name[:31] or "Table"


def clamp_box(
    box: tuple[float, float, float, float],
    width: int,
    height: int,
    padding: int = 0,
     ) -> tuple[int, int, int, int]:
    x1, y1, x2, y2 = box

    return (
        max(0, int(round(x1)) - padding),
        max(0, int(round(y1)) - padding),
        min(width, int(round(x2)) + padding),
        min(height, int(round(y2)) + padding),
    )


def box_area(box: tuple[float, float, float, float]) -> float:
    x1, y1, x2, y2 = box
    return max(0.0, x2 - x1) * max(0.0, y2 - y1)


def intersection_area(
    box_a: tuple[float, float, float, float],
    box_b: tuple[float, float, float, float],
    ) -> float:
    ax1, ay1, ax2, ay2 = box_a
    bx1, by1, bx2, by2 = box_b

    x1 = max(ax1, bx1)
    y1 = max(ay1, by1)
    x2 = min(ax2, bx2)
    y2 = min(ay2, by2)

    return max(0.0, x2 - x1) * max(0.0, y2 - y1)


def box_iou(
    box_a: tuple[float, float, float, float],
    box_b: tuple[float, float, float, float],
    ) -> float:
    intersection = intersection_area(box_a, box_b)
    union = box_area(box_a) + box_area(box_b) - intersection
    return intersection / union if union > 0 else 0.0


def remove_duplicate_detections(
    detections: list[Detection],
    iou_threshold: float,
    ) -> list[Detection]:
    """
    Remove near-identical model boxes while retaining the highest-confidence box.
    """
    kept: list[Detection] = []

    for item in sorted(detections, key=lambda detection: detection.score, reverse=True):
        duplicate = any(
            item.label == old.label
            and box_iou(item.box, old.box) >= iou_threshold
            for old in kept
        )
        if not duplicate:
            kept.append(item)

    return kept


def axis_length(
    detection: Detection,
    axis: str,
) -> float:
    if axis == "row":
        return max(0.0, detection.box[3] - detection.box[1])
    return max(0.0, detection.box[2] - detection.box[0])


def axis_center(
    detection: Detection,
    axis: str,
) -> float:
    if axis == "row":
        return (detection.box[1] + detection.box[3]) / 2
    return (detection.box[0] + detection.box[2]) / 2


def orthogonal_overlap_ratio(
    first: Detection,
    second: Detection,
    axis: str,
) -> float:
    if axis == "row":
        start = max(first.box[0], second.box[0])
        end = min(first.box[2], second.box[2])
        base = min(first.box[2] - first.box[0], second.box[2] - second.box[0])
    else:
        start = max(first.box[1], second.box[1])
        end = min(first.box[3], second.box[3])
        base = min(first.box[3] - first.box[1], second.box[3] - second.box[1])

    if base <= 0:
        return 0.0
    return max(0.0, end - start) / base


def collapse_nearby_structure_detections(
    detections: list[Detection],
    axis: str,
    center_tolerance_ratio: float = 0.4,
    overlap_ratio_threshold: float = 0.9,
) -> list[Detection]:
    """
    Remove nearly identical row/column detections that survive IoU-based filtering.
    """
    kept: list[Detection] = []

    for item in sorted(detections, key=lambda detection: detection.score, reverse=True):
        item_length = axis_length(item, axis)
        duplicate = False

        for existing in kept:
            existing_length = axis_length(existing, axis)
            tolerance = max(
                3.0,
                min(item_length, existing_length) * center_tolerance_ratio,
            )
            centers_are_close = (
                abs(axis_center(item, axis) - axis_center(existing, axis))
                <= tolerance
            )
            overlap_ratio = orthogonal_overlap_ratio(item, existing, axis)

            if centers_are_close and overlap_ratio >= overlap_ratio_threshold:
                duplicate = True
                break

        if not duplicate:
            kept.append(item)

    return kept


# =============================================================================
# PDF RENDERING
# =============================================================================

def render_pdf_pages(pdf_path: Path, dpi: int) -> list[Image.Image]:
    """
    Render all PDF pages to RGB PIL images.

    A higher DPI improves small-text OCR and structure recognition, but increases
    processing time and memory usage.
    """
    images: list[Image.Image] = []
    zoom = dpi / 72.0
    matrix = fitz.Matrix(zoom, zoom)

    with fitz.open(pdf_path) as document:
        for page in document:
            pixmap = page.get_pixmap(matrix=matrix, alpha=False)
            image = Image.frombytes(
                "RGB",
                (pixmap.width, pixmap.height),
                pixmap.samples,
            )
            images.append(image)

    return images


# =============================================================================
# TABLE TRANSFORMER
# =============================================================================

class TableTransformerEngine:
    def __init__(self) -> None:
        self.device = torch.device(
            "cuda" if torch.cuda.is_available() else "cpu"
        )
        # HuggingFace model inference is not safely concurrent across threads.
        self._inference_lock = Lock()

        if self.device.type == "cuda":
            torch.backends.cudnn.benchmark = True

        print(f"Using device: {self.device}")
        print(f"OCR engine: {OCR_ENGINE}")
        print(f"OCR workers: {effective_ocr_workers()}")

        self.detection_processor = AutoImageProcessor.from_pretrained(
            DETECTION_MODEL_NAME
        )
        # Avoid meta-tensor loads (accelerate/low_cpu_mem_usage) that break .to().
        self.detection_model = (
            TableTransformerForObjectDetection.from_pretrained(
                DETECTION_MODEL_NAME,
                low_cpu_mem_usage=False,
            )
            .to(self.device)
            .eval()
        )

        self.structure_processor = AutoImageProcessor.from_pretrained(
            STRUCTURE_MODEL_NAME
        )
        # transformers>=4.4x expects both shortest_edge and longest_edge.
        # v1.1 structure configs only ship longest_edge and crash on resize.
        self._normalize_processor_size(self.detection_processor)
        self._normalize_processor_size(self.structure_processor)

        self.structure_model = (
            TableTransformerForObjectDetection.from_pretrained(
                STRUCTURE_MODEL_NAME,
                low_cpu_mem_usage=False,
            )
            .to(self.device)
            .eval()
        )

    @staticmethod
    def _normalize_processor_size(processor) -> None:
        size = getattr(processor, "size", None)
        if not isinstance(size, dict):
            return

        if "shortest_edge" in size and "longest_edge" in size:
            return
        if "height" in size and "width" in size:
            return

        longest = int(size.get("longest_edge") or size.get("max_size") or 800)
        shortest = int(size.get("shortest_edge") or min(800, longest))
        processor.size = {
            "shortest_edge": shortest,
            "longest_edge": longest,
        }

    @staticmethod
    def _decode_results(
        result: dict,
        id2label: dict,
    ) -> list[Detection]:
        decoded: list[Detection] = []

        for score, label_tensor, box_tensor in zip(
            result["scores"],
            result["labels"],
            result["boxes"],
        ):
            label_id = int(label_tensor.item())
            label = id2label.get(
                label_id,
                id2label.get(str(label_id), str(label_id)),
            )

            decoded.append(
                Detection(
                    label=str(label).lower(),
                    score=float(score.item()),
                    box=tuple(
                        float(coordinate)
                        for coordinate in box_tensor.tolist()
                    ),
                )
            )

        return decoded

    @staticmethod
    def _move_inputs_to_device(inputs: dict, device: torch.device) -> dict:
        non_blocking = device.type == "cuda"
        return {
            key: value.to(device, non_blocking=non_blocking)
            for key, value in inputs.items()
        }

    def detect_tables(self, page_image: Image.Image) -> list[Detection]:
        with self._inference_lock:
            inputs = self.detection_processor(
                images=page_image,
                return_tensors="pt",
            )
            inputs = self._move_inputs_to_device(inputs, self.device)

            with torch.inference_mode():
                outputs = self.detection_model(**inputs)

            target_sizes = torch.tensor(
                [[page_image.height, page_image.width]],
                device=self.device,
            )

            result = self.detection_processor.post_process_object_detection(
                outputs,
                threshold=DETECTION_THRESHOLD,
                target_sizes=target_sizes,
            )[0]

            detections = self._decode_results(
                result,
                self.detection_model.config.id2label,
            )

        detections = [
            item
            for item in detections
            if item.label in {"table", "table rotated"}
        ]

        return remove_duplicate_detections(
            detections,
            iou_threshold=0.70,
        )

    def recognize_structure(
        self,
        table_image: Image.Image,
    ) -> list[Detection]:
        with self._inference_lock:
            inputs = self.structure_processor(
                images=table_image,
                return_tensors="pt",
            )
            inputs = self._move_inputs_to_device(inputs, self.device)

            with torch.inference_mode():
                outputs = self.structure_model(**inputs)

            target_sizes = torch.tensor(
                [[table_image.height, table_image.width]],
                device=self.device,
            )

            result = self.structure_processor.post_process_object_detection(
                outputs,
                threshold=STRUCTURE_THRESHOLD,
                target_sizes=target_sizes,
            )[0]

            return self._decode_results(
                result,
                self.structure_model.config.id2label,
            )


# =============================================================================
# OCR ADAPTER
# =============================================================================

def cell_ink_stats(
    cell_image: Image.Image,
    dark_threshold: int = 170,
) -> tuple[float, int, bool]:
    """
    Return (ink_ratio, dark_pixel_count, looks_like_grid_line_only).

    Grid-line-only crops are treated as blank even if they have dark pixels.
    Thin real glyphs like "1" are NOT treated as grid lines.
    """
    gray = cell_image.convert("L")
    width, height = gray.size
    if width < 2 or height < 2:
        return 0.0, 0, True

    inset_x = max(1, min(2, width // 12))
    inset_y = max(1, min(2, height // 12))
    if width > 2 * inset_x and height > 2 * inset_y:
        gray = gray.crop(
            (inset_x, inset_y, width - inset_x, height - inset_y)
        )

    pixels = np.asarray(gray, dtype=np.uint8)
    inner_h, inner_w = pixels.shape
    total_pixels = pixels.size or 1
    dark = pixels < dark_threshold
    dark_pixels = int(dark.sum())
    ink_ratio = dark_pixels / total_pixels

    if dark_pixels == 0:
        return 0.0, 0, False

    col_counts = dark.sum(axis=0)
    row_counts = dark.sum(axis=1)

    # Border-band test: true grid lines sit on the cell edge, not the center.
    x_band = max(1, inner_w // 8)
    y_band = max(1, inner_h // 8)

    left_dark = int(col_counts[:x_band].sum())
    right_dark = int(col_counts[-x_band:].sum())
    center_col_dark = (
        int(col_counts[x_band:-x_band].sum()) if inner_w > 2 * x_band else 0
    )
    top_dark = int(row_counts[:y_band].sum())
    bottom_dark = int(row_counts[-y_band:].sum())
    center_row_dark = (
        int(row_counts[y_band:-y_band].sum()) if inner_h > 2 * y_band else 0
    )

    looks_like_vertical_line = (
        dark_pixels >= max(8, inner_h // 4)
        and (left_dark + right_dark) >= int(0.80 * dark_pixels)
        and center_col_dark <= int(0.20 * dark_pixels)
    )
    looks_like_horizontal_line = (
        dark_pixels >= max(8, inner_w // 4)
        and (top_dark + bottom_dark) >= int(0.80 * dark_pixels)
        and center_row_dark <= int(0.20 * dark_pixels)
    )
    looks_like_grid_line = looks_like_vertical_line or looks_like_horizontal_line
    return ink_ratio, dark_pixels, looks_like_grid_line


def cell_has_visible_ink(
    cell_image: Image.Image,
    dark_threshold: int = 170,
    minimum_ink_ratio: float = 0.006,
    minimum_dark_pixels: int = 8,
) -> bool:
    """
    Return True when a cell crop contains enough dark pixels to justify OCR.

    Ignores a thin border so table grid lines alone do not count as content.
    """
    ink_ratio, dark_pixels, looks_like_grid_line = cell_ink_stats(
        cell_image,
        dark_threshold=dark_threshold,
    )
    if looks_like_grid_line:
        return False
    if dark_pixels < minimum_dark_pixels:
        return False
    return ink_ratio >= minimum_ink_ratio


def cell_is_definitely_blank(cell_image: Image.Image) -> bool:
    """Strict blank check used to skip OCR / clear invented text."""
    ink_ratio, dark_pixels, looks_like_grid_line = cell_ink_stats(cell_image)
    if looks_like_grid_line:
        return True
    return dark_pixels < 4 or ink_ratio < 0.002


def ocr_with_confidence(
    image: Image.Image,
    config: str,
    ) -> tuple[str, float]:
    """Run Tesseract and return recognized text plus mean word confidence."""
    data = pytesseract.image_to_data(
        image,
        config=config,
        output_type=pytesseract.Output.DICT,
    )

    texts: list[str] = []
    confidences: list[float] = []

    for text, confidence in zip(data.get("text", []), data.get("conf", [])):
        cleaned = str(text or "").strip()
        try:
            confidence_value = float(confidence)
        except (TypeError, ValueError):
            continue

        if not cleaned or confidence_value < 0:
            continue

        texts.append(cleaned)
        confidences.append(confidence_value)

    if not confidences:
        return "", 0.0

    return " ".join(texts), sum(confidences) / len(confidences)


def is_gibberish_token(token: str) -> bool:
    """Heuristic check for OCR hallucinated alphabetic tokens."""
    word = token.lower()
    if not word.isalpha():
        return False

    protected = {
        "and", "the", "for", "not", "yes", "nil", "ltd", "pvt", "huf",
        "mr", "mrs", "ms", "dr", "no", "sr", "qty", "nos", "amt", "ref",
        "date", "name", "code", "type", "unit", "rate", "total", "page",
        "from", "with", "this", "that", "each", "paid", "free", "none",
        "cash", "bank", "bill", "item", "amount", "number", "invoice",
        "address", "details", "description", "particulars", "quantity",
        "value", "price", "weight", "remarks", "month", "year", "days",
    }
    if word in protected:
        return False

    # Keep common uppercase-looking codes/acronyms (INR, HSN, GST...).
    if token.isupper() and 2 <= len(token) <= 5:
        return False

    # Keep normal short Title Case tokens (names/labels); only reject tiny crumbs.
    if len(word) <= 2:
        return True

    vowels = sum(character in "aeiou" for character in word)
    if len(word) >= 4 and vowels == 0:
        return True

    # Very low vowel density is typical of OCR junk.
    if len(word) >= 5 and vowels / len(word) < 0.2:
        return True

    # Repeated/near-repeated characters: "taee", "aaaa", "llll".
    if len(word) >= 4 and len(set(word)) <= 2:
        return True

    # Rare-letter heavy fragments are usually noise.
    rare = sum(character in "qzxjvwk" for character in word)
    if len(word) >= 4 and rare / len(word) >= 0.4:
        return True

    # Short random-looking words with awkward consonant clusters.
    if len(word) <= 6 and re.search(r"[bcdfghjklmnpqrstvwxyz]{4,}", word):
        return True

    # Weird mixed casing (e.g. "iRae", "aFeE") is typical OCR junk.
    if (
        len(token) >= 4
        and not token.islower()
        and not token.isupper()
        and not token.istitle()
        and len(token) <= 8
    ):
        return True

    # Short all-lowercase unknown tokens are common blank-cell hallucinations.
    if token.islower() and 3 <= len(token) <= 6 and word not in protected:
        return True

    return False


def tesseract_ocr(cell_image: Image.Image) -> str:
    """
    Cell OCR using Tesseract (pytesseract).

    Accepts a PIL.Image.Image and returns recognized text as a string.
    """
    if cell_image.width < 2 or cell_image.height < 2:
        return ""

    # Blank / near-blank cells: Tesseract invents random strings from noise.
    if not cell_has_visible_ink(cell_image):
        return ""

    enlarged = cell_image.resize(
        (cell_image.width * 2, cell_image.height * 2),
        Image.Resampling.LANCZOS,
    ).convert("L")
    contrasted = ImageOps.autocontrast(enlarged)
    thresholded = contrasted.point(
        lambda pixel: 255 if pixel > 180 else 0,
        mode="1",
    ).convert("L")

    def normalize_text(text: str) -> str:
        normalized_lines = [
            " ".join(line.split())
            for line in text.splitlines()
            if line.strip()
        ]
        return "\n".join(normalized_lines).strip()

    def score_text(text: str) -> tuple[int, int, int]:
        cleaned = cleanup_extracted_cell_text(text)
        if not cleaned:
            return (-1, -1, -1)

        # Never prefer OCR junk over an empty cell.
        if not has_meaningful_cell_content(cleaned):
            return (-1, -1, -1)

        alnum_count = sum(character.isalnum() for character in cleaned)
        length_score = len(cleaned.replace(" ", ""))
        return (10, alnum_count, length_score)

    best_text = ""
    best_score = (-1, -1, -1)
    best_confidence = -1.0

    # Same variants/PSMs as before, but stop early once a strong result is found.
    for image_variant in (contrasted, thresholded):
        for psm in (6, 7, 8):
            config = f"--oem 3 --psm {psm}"
            text, mean_confidence = ocr_with_confidence(
                image_variant,
                config=config,
            )
            normalized = normalize_text(text)
            cleaned = cleanup_extracted_cell_text(normalized)
            if not cleaned:
                continue

            # Digits/dates and placeholders can be valid at lower confidence.
            if is_legitimate_placeholder(cleaned):
                min_confidence = 25.0
            elif re.search(r"\d", cleaned):
                min_confidence = 35.0
            else:
                min_confidence = 42.0
            if mean_confidence < min_confidence:
                continue

            score = score_text(normalized)
            if score > best_score:
                best_score = score
                best_text = normalized
                best_confidence = mean_confidence

                if best_confidence >= OCR_EARLY_EXIT_CONFIDENCE:
                    return best_text

    return best_text


_easyocr_reader = None
_easyocr_lock = Lock()


def get_easyocr_reader():
    """Lazy-load a shared EasyOCR reader (expensive to create)."""
    global _easyocr_reader
    if _easyocr_reader is None:
        with _easyocr_lock:
            if _easyocr_reader is None:
                try:
                    import easyocr
                except ModuleNotFoundError as exc:
                    raise ModuleNotFoundError(
                        "EasyOCR is not installed. Run: pip install easyocr"
                    ) from exc

                use_gpu = torch.cuda.is_available()
                print(f"Loading EasyOCR reader (gpu={use_gpu})...")
                _easyocr_reader = easyocr.Reader(["en"], gpu=use_gpu)
    return _easyocr_reader


def easyocr_ocr(cell_image: Image.Image) -> str:
    """
    Cell OCR using EasyOCR (same style as OCR/EOCR.py).

    Accepts a PIL.Image.Image and returns recognized text as a string.
    """
    if cell_image.width < 2 or cell_image.height < 2:
        return ""

    if not cell_has_visible_ink(cell_image):
        return ""

    enlarged = cell_image.resize(
        (
            max(2, cell_image.width * 2),
            max(2, cell_image.height * 2),
        ),
        Image.Resampling.LANCZOS,
    ).convert("RGB")
    img_array = np.array(enlarged)

    reader = get_easyocr_reader()
    with _easyocr_lock:
        results = reader.readtext(img_array, detail=0, paragraph=False)

    lines = [
        " ".join(str(line).split())
        for line in results
        if str(line).strip()
    ]
    return "\n".join(lines).strip()


def resolve_ocr_function() -> Callable[[Image.Image], str]:
    """Return the OCR function selected by OCR_ENGINE."""
    engine = str(OCR_ENGINE or "tesseract").strip().lower()
    if engine == "tesseract":
        return tesseract_ocr
    if engine == "easyocr":
        return easyocr_ocr
    raise ValueError(
        f'Unsupported OCR_ENGINE={OCR_ENGINE!r}. '
        'Use "tesseract" or "easyocr".'
    )


# Backward-compatible alias used by extract_tables_from_pdf default arg.
default_ocr = tesseract_ocr


LEGITIMATE_CELL_PLACEHOLDERS = {
    "-",
    "–",
    "—",
    "nil",
    "n/a",
    "na",
    "none",
    "null",
}


def is_legitimate_placeholder(value: str | None) -> bool:
    text = str(value or "").strip().lower()
    return text in LEGITIMATE_CELL_PLACEHOLDERS


def is_probable_ocr_noise(line: str) -> bool:
    """Identify OCR fragments without removing normal cell values."""
    if is_legitimate_placeholder(line):
        return False

    tokens = re.findall(r"[A-Za-z0-9]+", line.lower())
    if not tokens:
        # Symbols-only cells such as "{" are not real table values.
        return bool(line.strip())

    if all(token.isdigit() for token in tokens):
        return False

    # Keep numeric-looking values that include separators/units crumbs.
    if re.fullmatch(r"[\d\s,./:%()\-A-Za-z]+", line) and re.search(r"\d", line):
        alpha_tokens = [token for token in tokens if not token.isdigit()]
        if not alpha_tokens or all(len(token) <= 3 for token in alpha_tokens):
            return False

    # Keep common short, valid forms such as "Mr. A. B.".
    if tokens[0] in {"mr", "mrs", "ms", "dr"}:
        return False

    original_tokens = re.findall(r"[A-Za-z0-9]+", line)
    alpha_original = [token for token in original_tokens if token.isalpha()]
    if (
        alpha_original
        and all(is_gibberish_token(token) for token in alpha_original)
        and not any(token.isdigit() for token in tokens)
    ):
        return True

    # Single short alphabetic token from a phantom/narrow column (e.g. ira, ode).
    if len(tokens) == 1:
        token = tokens[0]
        if token.isdigit():
            return False
        if len(token) <= 2:
            return True
        return is_gibberish_token(original_tokens[0] if original_tokens else token)

    if len(tokens) < 3:
        # Two short leftovers such as "fe i" / "ae fe".
        if all(len(token) <= 2 for token in tokens):
            return True
        return all(
            (not token.isdigit()) and is_gibberish_token(original)
            for token, original in zip(tokens, original_tokens)
        )

    lengths = [len(token) for token in tokens]
    if max(lengths) <= 2:
        return True

    # Examples: "i. tae ee" and similar short, disconnected fragments.
    protected_words = {"and", "the", "for", "not", "yes", "nil"}
    if alpha_original and sum(
        is_gibberish_token(token) for token in alpha_original
    ) >= max(1, len(alpha_original) - 1):
        return True

    return (
        max(lengths) <= 3
        and sum(length <= 2 for length in lengths) >= 2
        and not any(token in protected_words for token in tokens)
    )


def remove_noise_symbol_patterns(text: str) -> str:
    """
    Remove continuous / patterned OCR noise symbols while keeping useful punctuation.

    Examples removed:
      "~—_ ."
      "~~__"
      "_ _ ."
      "...."

    Examples kept:
      "12.50"
      "10%"
      "2024-05-01"
      "A-1"
      "-"
      "Nil"
    """
    if not text:
        return ""

    stripped = text.strip()
    if is_legitimate_placeholder(stripped):
        return "-" if stripped in {"-", "–", "—"} else stripped

    noise_chars = r"~`^¬¦|•·…—–_=*#@∞¤§±\"'‹›«»˚¨´"
    noise_or_filler = rf"[{noise_chars}.\-]"

    # Whole-cell / whole-line is only decorative symbols (but not a lone "-").
    if (
        re.fullmatch(rf"[\s{noise_chars}.\-]+", text)
        and not re.search(r"[A-Za-z0-9]", text)
        and not is_legitimate_placeholder(stripped)
    ):
        return ""

    # Continuous noise run: "~—_", "___", "~~~"
    text = re.sub(rf"[{noise_chars}]{{2,}}", " ", text)

    # Repeated same filler punctuation: "....", "----", "===="
    text = re.sub(r"([.\-_=~—–])\1+", " ", text)

    # Spaced symbol patterns: "~ — _ ." or "_ . ~ —"
    text = re.sub(
        rf"(?:(?<=\s)|^){noise_or_filler}(?:\s+{noise_or_filler}){{1,}}(?=\s|$)",
        " ",
        text,
    )

    cleaned_tokens: list[str] = []
    for token in text.split():
        if is_legitimate_placeholder(token):
            cleaned_tokens.append("-" if token in {"-", "–", "—"} else token)
            continue

        # Drop tokens that are only noise/filler symbols.
        if re.fullmatch(rf"[{noise_chars}.\-]+", token):
            # Keep a lone meaningful unit marker if it appears alone.
            if token == "%":
                cleaned_tokens.append(token)
            continue

        # Strip leading/trailing decorative symbols, keep inner useful punctuation.
        token = re.sub(rf"^[{noise_chars}]+", "", token)
        token = re.sub(rf"[{noise_chars}]+$", "", token)

        # Remove internal continuous noise chunks inside a mixed token.
        token = re.sub(rf"[{noise_chars}]{{2,}}", "", token)

        if not token:
            continue
        if re.fullmatch(rf"[{noise_chars}.\-]+", token):
            continue

        cleaned_tokens.append(token)

    return " ".join(cleaned_tokens).strip()


def cleanup_extracted_cell_text(value: str | None) -> str:
    """Remove OCR-only symbols and clearly meaningless text lines."""
    text = str(value or "")
    if is_legitimate_placeholder(text.strip()):
        return "-" if text.strip() in {"-", "–", "—"} else text.strip()

    # Vertical grid lines are OCR artefacts, not table content.
    text = re.sub(r"[|¦\u2502\u2503\u2551\u254E\u254F]+", " ", text)

    kept_lines = []
    for line in text.splitlines():
        normalized = " ".join(line.split())
        normalized = remove_noise_symbol_patterns(normalized)
        if not normalized:
            continue
        if is_legitimate_placeholder(normalized) or not is_probable_ocr_noise(
            normalized
        ):
            kept_lines.append(normalized)

    return "\n".join(kept_lines).strip()


def has_meaningful_cell_content(value: str | None) -> bool:
    """Return whether a cleaned cell contains real table content."""
    if is_legitimate_placeholder(value):
        return True

    # Use lightweight cleanup here to avoid recursion with cleanup_extracted_cell_text.
    text = str(value or "")
    text = re.sub(r"[|¦\u2502\u2503\u2551\u254E\u254F]+", " ", text)
    text = "\n".join(
        remove_noise_symbol_patterns(" ".join(line.split()))
        for line in text.splitlines()
        if line.strip()
    )
    text = "\n".join(
        line
        for line in text.splitlines()
        if line.strip()
        and (
            is_legitimate_placeholder(line.strip())
            or not is_probable_ocr_noise(line.strip())
        )
    ).strip()
    if not text:
        return False

    if is_legitimate_placeholder(text):
        return True

    words = re.findall(r"[A-Za-z]+", text)
    real_words = [
        word
        for word in words
        if len(word) >= 3 and not is_gibberish_token(word)
    ]
    # Prefer real words; ignore hallucinated OCR crumbs.
    if real_words:
        return True

    # Allow uppercase codes such as INR / USA / HUF.
    if any(len(word) == 3 and word.isupper() for word in words):
        return True

    protected = {
        "and", "the", "for", "not", "yes", "nil", "ltd", "pvt", "huf",
        "mr", "mrs", "ms", "dr", "no", "sr", "qty", "nos", "amt", "ref",
    }
    if any(word.lower() in protected for word in words):
        return True

    # Do not discard a valid date, amount, percentage, or other numeric field
    # merely because its column header was missed by OCR.
    return bool(
        re.search(r"\d", text)
        and re.fullmatch(r"[\d\s,./:%()\-A-Za-z]+", text)
    )


def _column_values(data: list[list[str]], column_index: int) -> list[str]:
    return [
        row[column_index] if column_index < len(row) else ""
        for row in data
    ]


def _cell_is_real_content(value: str | None) -> bool:
    """True for real text or legitimate placeholders like '-', 'Nil'."""
    text = str(value or "").strip()
    if not text:
        return False
    if is_legitimate_placeholder(text):
        return True
    return has_meaningful_cell_content(text)


def _cell_is_noise_or_null(value: str | None) -> bool:
    """True for empty/null cells or OCR noise fragments."""
    text = str(value or "").strip()
    if not text:
        return True
    if is_legitimate_placeholder(text):
        return False
    return not has_meaningful_cell_content(text)


def _header_is_noisy(header: str | None) -> bool:
    """Empty headers are not noisy; only gibberish/OCR crumbs count."""
    text = str(header or "").strip()
    if not text:
        return False
    if is_legitimate_placeholder(text):
        return False
    if has_meaningful_cell_content(text):
        return False
    return True


def is_junk_generated_column(values: list[str]) -> bool:
    """
    Drop a column only when its header is noisy AND every cell is noise/null.

    Sparse real columns (many empty cells, a few real values, or mostly
    '-' / 'Nil') must always be kept.
    """
    if not values:
        return True

    # Any real content or placeholder anywhere → keep.
    if any(_cell_is_real_content(cell) for cell in values):
        return False

    header = values[0]
    if not _header_is_noisy(header):
        # Empty/quiet header with empty or sparse body → keep.
        return False

    # Noisy header and nothing but noise/null below → drop.
    return all(_cell_is_noise_or_null(cell) for cell in values)


def prune_generated_columns(
    data: list[list[str]],
    merge_ranges: list[tuple[int, int, int, int]],
    ) -> tuple[list[list[str]], list[tuple[int, int, int, int]]]:
    """Drop columns only when header is noisy and all cells are noise/null."""
    if not data or not data[0]:
        return data, merge_ranges

    column_count = len(data[0])
    if column_count <= 1:
        return data, merge_ranges

    keep_indices = [0]
    for column_index in range(1, column_count):
        values = _column_values(data, column_index)
        if is_junk_generated_column(values):
            continue
        keep_indices.append(column_index)

    if len(keep_indices) == column_count:
        return data, merge_ranges

    index_map = {
        old_index: new_index
        for new_index, old_index in enumerate(keep_indices)
    }
    new_data = [
        [row[index] if index < len(row) else "" for index in keep_indices]
        for row in data
    ]

    new_merges: list[tuple[int, int, int, int]] = []
    for row_start, row_end, column_start, column_end in merge_ranges:
        mapped_columns = [
            index_map[index]
            for index in range(column_start, column_end + 1)
            if index in index_map
        ]
        if not mapped_columns:
            continue

        if (
            column_start not in index_map
            and row_start < len(data)
            and column_start < len(data[row_start])
        ):
            new_data[row_start][mapped_columns[0]] = data[row_start][column_start]

        new_merges.append(
            (
                row_start,
                row_end,
                min(mapped_columns),
                max(mapped_columns),
            )
        )

    print(
        f"  Removed {column_count - len(keep_indices)} noisy junk column(s)."
    )
    return new_data, new_merges


def rows_look_like_duplicates(
    first_row: list[str],
    second_row: list[str],
) -> bool:
    comparable_pairs = 0
    matching_pairs = 0
    meaningful_first = 0
    meaningful_second = 0

    for first_cell, second_cell in zip(first_row, second_row, strict=False):
        first_text = cleanup_extracted_cell_text(first_cell)
        second_text = cleanup_extracted_cell_text(second_cell)

        first_meaningful = has_meaningful_cell_content(first_text)
        second_meaningful = has_meaningful_cell_content(second_text)
        meaningful_first += int(first_meaningful)
        meaningful_second += int(second_meaningful)

        if not first_meaningful and not second_meaningful:
            continue

        comparable_pairs += 1
        if first_text == second_text:
            matching_pairs += 1
            continue

        if first_text and second_text and (
            first_text in second_text or second_text in first_text
        ):
            matching_pairs += 1

    minimum_meaningful = min(meaningful_first, meaningful_second)
    if minimum_meaningful < 2 or comparable_pairs == 0:
        return False

    coverage = matching_pairs / comparable_pairs
    return coverage >= 0.75


def merge_similar_rows(
    first_row: list[str],
    second_row: list[str],
) -> list[str]:
    merged: list[str] = []
    column_count = max(len(first_row), len(second_row))

    for index in range(column_count):
        first_cell = first_row[index] if index < len(first_row) else ""
        second_cell = second_row[index] if index < len(second_row) else ""
        first_text = cleanup_extracted_cell_text(first_cell)
        second_text = cleanup_extracted_cell_text(second_cell)

        if not first_text:
            merged.append(second_text)
            continue
        if not second_text:
            merged.append(first_text)
            continue

        first_meaningful = has_meaningful_cell_content(first_text)
        second_meaningful = has_meaningful_cell_content(second_text)

        if first_meaningful and not second_meaningful:
            merged.append(first_text)
        elif second_meaningful and not first_meaningful:
            merged.append(second_text)
        elif len(second_text) > len(first_text) and first_text in second_text:
            merged.append(second_text)
        else:
            merged.append(first_text)

    return merged


def deduplicate_adjacent_rows(
    data: list[list[str]],
    merge_ranges: list[tuple[int, int, int, int]],
) -> tuple[list[list[str]], list[tuple[int, int, int, int]]]:
    if len(data) < 2:
        return data, merge_ranges

    new_data: list[list[str]] = []
    index_map: dict[int, int] = {}
    source_index = 0

    while source_index < len(data):
        current_row = data[source_index]

        if (
            source_index + 1 < len(data)
            and rows_look_like_duplicates(current_row, data[source_index + 1])
        ):
            current_row = merge_similar_rows(current_row, data[source_index + 1])
            index_map[source_index] = len(new_data)
            index_map[source_index + 1] = len(new_data)
            new_data.append(current_row)
            source_index += 2
            continue

        index_map[source_index] = len(new_data)
        new_data.append(current_row)
        source_index += 1

    if len(new_data) == len(data):
        return data, merge_ranges

    new_merges: list[tuple[int, int, int, int]] = []
    for row_start, row_end, column_start, column_end in merge_ranges:
        mapped_rows = [
            index_map[index]
            for index in range(row_start, row_end + 1)
            if index in index_map
        ]
        if not mapped_rows:
            continue

        new_merges.append(
            (
                min(mapped_rows),
                max(mapped_rows),
                column_start,
                column_end,
            )
        )

    print(f"  Collapsed {len(data) - len(new_data)} duplicate/noisy row(s).")
    return new_data, new_merges


# Example adapter for your existing OCR:
#
# def my_existing_ocr(cell_image: Image.Image) -> str:
#     result = your_ocr_model.recognize(cell_image)
#     return result["text"].strip()
#
# Then change the last function call to:
#
# extract_tables_from_pdf(
#     INPUT_PDF,
#     OUTPUT_XLSX,
#     ocr_function=my_existing_ocr,
# )


# =============================================================================
# TABLE STRUCTURE TO GRID
# =============================================================================

def prepare_rows_columns_and_spans(
    structure: list[Detection],
    ) -> tuple[list[Detection], list[Detection], list[Detection]]:
    rows = [
        item
        for item in structure
        if item.label == "table row"
    ]
    columns = [
        item
        for item in structure
        if item.label == "table column"
    ]
    spans = [
        item
        for item in structure
        if item.label == "table spanning cell"
    ]

    rows = remove_duplicate_detections(rows, iou_threshold=0.85)
    columns = remove_duplicate_detections(columns, iou_threshold=0.85)
    spans = remove_duplicate_detections(spans, iou_threshold=0.85)

    rows = collapse_nearby_structure_detections(rows, axis="row")
    columns = collapse_nearby_structure_detections(columns, axis="column")

    rows.sort(key=lambda item: (item.box[1] + item.box[3]) / 2)
    columns.sort(key=lambda item: (item.box[0] + item.box[2]) / 2)

    return rows, columns, spans


def grid_cell_box(
    row: Detection,
    column: Detection,
    image_width: int,
    image_height: int,
    ) -> tuple[int, int, int, int]:
    """
    Build a cell using the intersection of one predicted row and one column.
    """
    x1 = max(0, int(round(column.box[0])))
    y1 = max(0, int(round(row.box[1])))
    x2 = min(image_width, int(round(column.box[2])))
    y2 = min(image_height, int(round(row.box[3])))

    if x2 <= x1 or y2 <= y1:
        return (0, 0, 0, 0)

    return (x1, y1, x2, y2)


def find_span_range(
    span: Detection,
    rows: list[Detection],
    columns: list[Detection],
    table_width: int,
    table_height: int,
    minimum_cell_overlap: float = 0.50,
    ) -> tuple[int, int, int, int] | None:
    """
    Convert a predicted spanning-cell box into zero-based grid coordinates.

    Returns:
        row_start, row_end, column_start, column_end
    """
    matched_rows: set[int] = set()
    matched_columns: set[int] = set()

    for row_index, row in enumerate(rows):
        for column_index, column in enumerate(columns):
            cell_box = grid_cell_box(
                row,
                column,
                table_width,
                table_height,
            )
            area = box_area(cell_box)

            if area <= 0:
                continue

            overlap_ratio = intersection_area(span.box, cell_box) / area

            if overlap_ratio >= minimum_cell_overlap:
                matched_rows.add(row_index)
                matched_columns.add(column_index)

    if not matched_rows or not matched_columns:
        return None

    row_start = min(matched_rows)
    row_end = max(matched_rows)
    column_start = min(matched_columns)
    column_end = max(matched_columns)

    if row_start == row_end and column_start == column_end:
        return None

    return row_start, row_end, column_start, column_end


def ranges_overlap(
    first: tuple[int, int, int, int],
    second: tuple[int, int, int, int],
    ) -> bool:
    first_row_start, first_row_end, first_col_start, first_col_end = first
    second_row_start, second_row_end, second_col_start, second_col_end = second

    rows_overlap = not (
        first_row_end < second_row_start
        or second_row_end < first_row_start
    )
    columns_overlap = not (
        first_col_end < second_col_start
        or second_col_end < first_col_start
    )

    return rows_overlap and columns_overlap


def recover_missed_cell_text(
    table_image: Image.Image,
    cell_box: tuple[int, int, int, int],
) -> str:
    """
    Targeted recovery for a single empty cell that still has real ink.

    Blank/grid-only cells stay blank; thin real values like "1" are recovered.
    """
    x1, y1, x2, y2 = cell_box
    if x2 <= x1 or y2 <= y1:
        return ""

    crop = table_image.crop((x1, y1, x2, y2))
    if cell_is_definitely_blank(crop):
        return ""
    if not cell_has_visible_ink(crop):
        # Try a tiny expansion only when the raw crop is borderline.
        expanded = clamp_box(
            (x1 - 2, y1 - 2, x2 + 2, y2 + 2),
            table_image.width,
            table_image.height,
        )
        crop = table_image.crop(expanded)
        if cell_is_definitely_blank(crop) or not cell_has_visible_ink(crop):
            return ""

    scale = 3 if min(crop.width, crop.height) < 70 else 2
    gray = crop.resize(
        (max(1, crop.width * scale), max(1, crop.height * scale)),
        Image.Resampling.LANCZOS,
    ).convert("L")
    gray = ImageOps.autocontrast(ImageOps.expand(gray, border=12, fill=255))

    best_text = ""
    best_key = (-1.0, -1, -1)

    for config in (
        "--oem 3 --psm 7",
        "--oem 3 --psm 8",
        "--oem 3 --psm 8 -c tessedit_char_whitelist=0123456789.,-/%:",
    ):
        text, confidence = ocr_with_confidence(gray, config=config)
        cleaned = cleanup_extracted_cell_text(text)
        if not cleaned or not has_meaningful_cell_content(cleaned):
            continue

        min_confidence = 38.0 if re.search(r"\d", cleaned) else 48.0
        if confidence < min_confidence:
            continue

        key = (
            confidence,
            1 if re.search(r"\d", cleaned) else 0,
            len(re.findall(r"[A-Za-z0-9]", cleaned)),
        )
        if key > best_key:
            best_key = key
            best_text = cleaned
            if confidence >= OCR_EARLY_EXIT_CONFIDENCE:
                return best_text

    return best_text


def _ocr_single_grid_cell(
    table_image: Image.Image,
    row: Detection,
    column: Detection,
    ocr_function: Callable[[Image.Image], str],
) -> str:
    """OCR one grid cell with the same retry/recovery logic as before."""
    raw_box = grid_cell_box(
        row,
        column,
        table_image.width,
        table_image.height,
    )

    if raw_box[2] <= raw_box[0] or raw_box[3] <= raw_box[1]:
        return ""

    x1, y1, x2, y2 = raw_box
    x1 = min(x2, x1 + CELL_INSET)
    y1 = min(y2, y1 + CELL_INSET)
    x2 = max(x1, x2 - CELL_INSET)
    y2 = max(y1, y2 - CELL_INSET)

    cell_image = table_image.crop((x1, y1, x2, y2))
    raw_image = table_image.crop(raw_box)

    # Only skip OCR when the cell is truly blank / border-only.
    if cell_is_definitely_blank(cell_image) and cell_is_definitely_blank(
        raw_image
    ):
        return ""

    cell_text = ""
    if cell_has_visible_ink(cell_image) or not cell_is_definitely_blank(
        cell_image
    ):
        cell_text = ocr_function(cell_image)

    if not cleanup_extracted_cell_text(cell_text):
        retry_box = clamp_box(
            (
                x1 - max(CELL_INSET, 2),
                y1 - max(CELL_INSET, 2),
                x2 + max(CELL_INSET, 2),
                y2 + max(CELL_INSET, 2),
            ),
            table_image.width,
            table_image.height,
        )
        retry_image = table_image.crop(retry_box)
        if not cell_is_definitely_blank(retry_image):
            retry_text = ocr_function(retry_image)
            if cleanup_extracted_cell_text(retry_text):
                cell_text = retry_text

        if not cleanup_extracted_cell_text(cell_text):
            if not cell_is_definitely_blank(raw_image):
                raw_text = ocr_function(raw_image)
                if cleanup_extracted_cell_text(raw_text):
                    cell_text = raw_text
                else:
                    cell_text = recover_missed_cell_text(
                        table_image,
                        raw_box,
                    )

    # Final guard: clear only definite blanks / non-meaningful junk.
    if cleanup_extracted_cell_text(cell_text):
        if cell_is_definitely_blank(cell_image) and cell_is_definitely_blank(
            raw_image
        ):
            return ""
        if not has_meaningful_cell_content(cell_text):
            return ""

    return cell_text


def extract_grid(
    table_image: Image.Image,
    rows: list[Detection],
    columns: list[Detection],
    spans: list[Detection],
    ocr_function: Callable[[Image.Image], str],
    ) -> tuple[
    list[list[str]],
    list[tuple[int, int, int, int]],
    ]:
    """
    OCR every ordinary cell, then replace predicted spanning ranges with one
    OCR result for the complete merged region.
    """
    data = [
        ["" for _ in columns]
        for _ in rows
    ]

    cell_jobs = [
        (row_index, column_index, row, column)
        for row_index, row in enumerate(rows)
        for column_index, column in enumerate(columns)
    ]

    def _run_job(
        job: tuple[int, int, Detection, Detection],
    ) -> tuple[int, int, str]:
        row_index, column_index, row, column = job
        text = _ocr_single_grid_cell(
            table_image,
            row,
            column,
            ocr_function,
        )
        return row_index, column_index, text

    workers = min(effective_ocr_workers(), len(cell_jobs))

    if len(cell_jobs) == 1 or workers <= 1:
        for job in cell_jobs:
            row_index, column_index, text = _run_job(job)
            data[row_index][column_index] = text
    else:
        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = [executor.submit(_run_job, job) for job in cell_jobs]
            for future in as_completed(futures):
                row_index, column_index, text = future.result()
                data[row_index][column_index] = text

    merge_ranges: list[tuple[int, int, int, int]] = []

    for span in sorted(
        spans,
        key=lambda item: item.score,
        reverse=True,
    ):
        merge_range = find_span_range(
            span,
            rows,
            columns,
            table_image.width,
            table_image.height,
        )

        if merge_range is None:
            continue

        if any(
            ranges_overlap(merge_range, existing)
            for existing in merge_ranges
        ):
            continue

        row_start, row_end, column_start, column_end = merge_range

        # Avoid false spans wiping distinct already-extracted middle cells.
        occupied_values = {
            cleanup_extracted_cell_text(data[row_index][column_index])
            for row_index in range(row_start, row_end + 1)
            for column_index in range(column_start, column_end + 1)
            if cleanup_extracted_cell_text(data[row_index][column_index])
        }
        if len(occupied_values) >= 2:
            continue

        x1 = min(
            columns[index].box[0]
            for index in range(column_start, column_end + 1)
        )
        x2 = max(
            columns[index].box[2]
            for index in range(column_start, column_end + 1)
        )
        y1 = min(
            rows[index].box[1]
            for index in range(row_start, row_end + 1)
        )
        y2 = max(
            rows[index].box[3]
            for index in range(row_start, row_end + 1)
        )

        merged_box = clamp_box(
            (x1, y1, x2, y2),
            table_image.width,
            table_image.height,
        )
        merged_image = table_image.crop(merged_box)
        merged_text = ocr_function(merged_image)
        cleaned_merged = cleanup_extracted_cell_text(merged_text)
        existing_anchor = cleanup_extracted_cell_text(
            data[row_start][column_start]
        )

        # Do not replace a good cell with empty/weaker merged OCR.
        if not cleaned_merged:
            continue
        if (
            existing_anchor
            and existing_anchor != cleaned_merged
            and existing_anchor not in cleaned_merged
            and len(existing_anchor) >= len(cleaned_merged)
        ):
            continue

        data[row_start][column_start] = merged_text

        for row_index in range(row_start, row_end + 1):
            for column_index in range(column_start, column_end + 1):
                if (
                    row_index != row_start
                    or column_index != column_start
                ):
                    data[row_index][column_index] = ""

        merge_ranges.append(merge_range)

    return data, merge_ranges


# =============================================================================
# DEBUG IMAGES
# =============================================================================

def save_structure_debug_image(
    table_image: Image.Image,
    rows: list[Detection],
    columns: list[Detection],
    spans: list[Detection],
    output_path: Path,
    ) -> None:
    """
    Red = rows, blue = columns, green = spanning cells.
    """
    debug_image = table_image.copy()
    draw = ImageDraw.Draw(debug_image)

    for row in rows:
        draw.rectangle(row.box, outline="red", width=2)

    for column in columns:
        draw.rectangle(column.box, outline="blue", width=2)

    for span in spans:
        draw.rectangle(span.box, outline="green", width=3)

    debug_image.save(output_path)


# =============================================================================
# EXCEL OUTPUT
# =============================================================================

def write_table_sheet(
    workbook: Workbook,
    sheet_name: str,
    data: list[list[str]],
    merge_ranges: list[tuple[int, int, int, int]],
     ) -> None:
    worksheet = workbook.create_sheet(
        title=safe_sheet_name(sheet_name)
    )
    worksheet.freeze_panes = "A2"

    for row_number, values in enumerate(data, start=1):
        for column_number, value in enumerate(values, start=1):
            cell = worksheet.cell(
                row=row_number,
                column=column_number,
                value=value,
            )
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )

    # Treat the first detected row as a probable header.
    if data:
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="D9EAF7",
            )

    # Determine widths before applying Excel merges.
    for column_number in range(1, len(data[0]) + 1 if data else 1):
        maximum_length = 0

        for row_number in range(1, len(data) + 1):
            value = worksheet.cell(
                row=row_number,
                column=column_number,
            ).value

            if value is None:
                continue

            maximum_length = max(
                maximum_length,
                max(
                    (
                        len(line)
                        for line in str(value).splitlines()
                    ),
                    default=0,
                ),
            )

        worksheet.column_dimensions[
            get_column_letter(column_number)
        ].width = min(max(maximum_length + 2, 10), 45)

    for row_start, row_end, column_start, column_end in merge_ranges:
        worksheet.merge_cells(
            start_row=row_start + 1,
            end_row=row_end + 1,
            start_column=column_start + 1,
            end_column=column_end + 1,
        )


# =============================================================================
# MAIN PIPELINE
# =============================================================================

def process_pdf_page(
    page_number: int,
    page_image: Image.Image,
    ocr_function: Callable[[Image.Image], str],
    engine: TableTransformerEngine,
) -> tuple[int, list[tuple[str, list[list[str]], list[tuple[int, int, int, int]]]]]:
    """
    Process one rendered PDF page and return extracted sheet payloads.

    Workbook writes happen on the main thread; worker threads only return data.
    Detection/structure share one engine (locked). OCR still runs per page.
    """
    tables = engine.detect_tables(page_image)

    print(
        f"Page {page_number}: "
        f"{len(tables)} table region(s) detected."
    )

    page_results: list[
        tuple[str, list[list[str]], list[tuple[int, int, int, int]]]
    ] = []

    for table_number, table in enumerate(tables, start=1):
        crop_box = clamp_box(
            table.box,
            page_image.width,
            page_image.height,
            padding=TABLE_PADDING,
        )
        table_image = page_image.crop(crop_box)

        if table.label == "table rotated":
            table_image = table_image.rotate(
                90,
                expand=True,
            )

        table_crop_path = (
            DEBUG_DIR
            / f"page_{page_number}_table_{table_number}.png"
        )
        DEBUG_DIR.mkdir(parents=True, exist_ok=True)
        table_image.save(table_crop_path)

        structure = engine.recognize_structure(table_image)
        rows, columns, spans = prepare_rows_columns_and_spans(
            structure
        )

        structure_debug_path = DEBUG_DIR / (
            f"page_{page_number}_table_{table_number}"
            "_structure.png"
        )
        DEBUG_DIR.mkdir(parents=True, exist_ok=True)
        save_structure_debug_image(
            table_image,
            rows,
            columns,
            spans,
            structure_debug_path,
        )

        if not rows or not columns:
            print(
                f"  Skipped page {page_number} table {table_number}: "
                "no usable row/column structure."
            )
            continue

        data, merge_ranges = extract_grid(
            table_image,
            rows,
            columns,
            spans,
            ocr_function,
        )
        data = [
            [cleanup_extracted_cell_text(value) for value in row]
            for row in data
        ]
        data, merge_ranges = prune_generated_columns(data, merge_ranges)
        data, merge_ranges = deduplicate_adjacent_rows(data, merge_ranges)

        sheet_name = f"P{page_number}_Table{table_number}"
        page_results.append((sheet_name, data, merge_ranges))

        print(
            f"  Extracted page {page_number} table {table_number}: "
            f"{len(rows)} row(s), {len(data[0]) if data else 0} column(s)."
        )

    return page_number, page_results

def extract_tables_from_pdf(
    input_pdf: Path,
    output_xlsx: Path,
    ocr_function: Callable[[Image.Image], str] = default_ocr,
    ) -> None:
    if not input_pdf.exists():
        raise FileNotFoundError(
            f"Input PDF was not found: {input_pdf.resolve()}"
        )

    DEBUG_DIR.mkdir(parents=True, exist_ok=True)

    # One shared engine avoids concurrent from_pretrained meta-tensor crashes.
    engine = TableTransformerEngine()
    pages = render_pdf_pages(input_pdf, PDF_DPI)

    workbook = Workbook()
    workbook.remove(workbook.active)

    extracted_table_count = 0
    page_payloads: list[
        tuple[int, list[tuple[str, list[list[str]], list[tuple[int, int, int, int]]]]]
    ] = []
    page_workers = min(effective_page_workers(), len(pages))

    if page_workers <= 1:
        for page_number, page_image in enumerate(pages, start=1):
            page_payloads.append(
                process_pdf_page(
                    page_number,
                    page_image,
                    ocr_function,
                    engine,
                )
            )
    else:
        print(f"Page workers: {page_workers}")
        with ThreadPoolExecutor(max_workers=page_workers) as executor:
            futures = [
                executor.submit(
                    process_pdf_page,
                    page_number,
                    page_image,
                    ocr_function,
                    engine,
                )
                for page_number, page_image in enumerate(pages, start=1)
            ]
            for future in as_completed(futures):
                page_payloads.append(future.result())

    for _page_number, page_results in sorted(page_payloads, key=lambda item: item[0]):
        for sheet_name, data, merge_ranges in page_results:
            write_table_sheet(
                workbook,
                sheet_name=sheet_name,
                data=data,
                merge_ranges=merge_ranges,
            )
            extracted_table_count += 1

    if extracted_table_count == 0:
        worksheet = workbook.create_sheet("No tables found")
        worksheet["A1"] = (
            "No table with usable row and column structure was detected."
        )

    workbook.save(output_xlsx)

    print()
    print(f"Tables extracted: {extracted_table_count}")
    print(f"Excel file: {output_xlsx.resolve()}")
    print(f"Debug images: {DEBUG_DIR.resolve()}")


if __name__ == "__main__":
    extract_tables_from_pdf(
        input_pdf=INPUT_PDF,
        output_xlsx=OUTPUT_XLSX,
        ocr_function=resolve_ocr_function(),
    )
