import pandas as pd
import re
from openpyxl import load_workbook


def add_gst_bill_columns(excel_input_path: str,
                        excel_output_path: str,
                        buyer_column_name: str = "GST Buyer",
                        seller_column_name: str = "GST Seller",
                        bill_no_column_name: str = "Bill_No",
                        bill_date_column_name: str = "Bill_Date") -> dict:
    # Load sheet as DataFrame to scan for patterns
    df = pd.read_excel(excel_input_path, header=None, dtype=str).fillna("")
    
    # Initialize variables
    buyer_gst = None
    seller_gst = None
    bill_no = None
    bill_date = None
    
    # GSTIN regex pattern
    gst_regex = re.compile(r"\b[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][1-9A-Z]Z[0-9A-Z]\b")
    
    # Bill No pattern (year-year /number format like "2024-2025 /457")
    bill_no_regex = re.compile(r"\b[0-9]{4}-[0-9]{4}\s*/\s*[0-9]+\b")
    
    # Date patterns (various formats)
    date_regex = re.compile(r"\b(?:[0-9]{1,2}[-/][0-9]{1,2}[-/][0-9]{4}|[0-9]{2}-[0-9]{2}-[0-9]{4})\b")
    
    # Scan all cells for patterns
    for i, row in df.iterrows():
        for j, cell in row.items():
            if not cell:
                continue
                
            cell_text = str(cell).replace('\n', ' ')
            
            # Look for GSTINs
            for gst in gst_regex.findall(cell_text):
                # Get context for buyer detection
                context_cells = row.tolist()
                if i > 0:
                    context_cells = df.iloc[i-1].tolist() + context_cells
                context_str = " ".join([str(c) for c in context_cells])
                
                # Check for buyer indicators
                if buyer_gst is None and re.search(
                    r"\b(Ms\.?|Mr\.?|M/s|Limited|Pvt\.?|LLP|India|Dist\.?|Tal\.?|PIN)\b",
                    context_str, flags=re.I):
                    buyer_gst = gst
                elif gst != buyer_gst and seller_gst is None:
                    seller_gst = gst
            
            # Look for Bill No pattern in same block as "BILL NO"
            if "BILL NO" in cell_text.upper():
                bill_matches = bill_no_regex.findall(cell_text)
                if bill_matches and bill_no is None:
                    bill_no = bill_matches[0].strip()
            
            # Look for Bill Date in same block as "BILL DATE"
            if "BILL DATE" in cell_text.upper():
                date_matches = date_regex.findall(cell_text)
                if date_matches and bill_date is None:
                    bill_date = date_matches[0].strip()
    
    # If no specific patterns found, try general search
    all_text = " ".join(df.values.flatten())
    
    # Backup searches if primary searches failed
    if not buyer_gst or not seller_gst:
        all_gsts = gst_regex.findall(all_text)
        for gst in all_gsts:
            if buyer_gst is None:
                buyer_gst = gst
            elif gst != buyer_gst and seller_gst is None:
                seller_gst = gst
    
    if bill_no is None:
        bill_matches = bill_no_regex.findall(all_text)
        if bill_matches:
            bill_no = bill_matches[0].strip()
    
    if bill_date is None:
        date_matches = date_regex.findall(all_text)
        if date_matches:
            bill_date = date_matches[0].strip()
    
    # Load workbook to preserve formatting
    wb = load_workbook(excel_input_path)
    ws = wb.active
    
    # Add headers in new columns (4 total columns)
    last_col = ws.max_column
    ws.cell(row=1, column=last_col+1, value=buyer_column_name)
    ws.cell(row=1, column=last_col+2, value=seller_column_name)
    ws.cell(row=1, column=last_col+3, value=bill_no_column_name)
    ws.cell(row=1, column=last_col+4, value=bill_date_column_name)
    
    # Fill values down to last row
    for row_idx in range(2, ws.max_row + 1):
        if buyer_gst:
            ws.cell(row=row_idx, column=last_col+1, value=buyer_gst)
        if seller_gst:
            ws.cell(row=row_idx, column=last_col+2, value=seller_gst)
        if bill_no:
            ws.cell(row=row_idx, column=last_col+3, value=bill_no)
        if bill_date:
            ws.cell(row=row_idx, column=last_col+4, value=bill_date)
    
    # Save updated workbook
    wb.save(excel_output_path)
    
    return {
        'buyer_gst': buyer_gst,
        'seller_gst': seller_gst,
        'bill_no': bill_no,
        'bill_date': bill_date
    }

add_gst_bill_columns(excel_input_path="extracted_tables.xlsx",
                    excel_output_path="extracted_tables_gst_bill.xlsx")