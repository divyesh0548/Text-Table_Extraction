from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Callable

import fitz  # PyMuPDF
import pytesseract
import torch
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw
from transformers import AutoImageProcessor, TableTransformerForObjectDetection


# =============================================================================
# CONFIGURATION
# =============================================================================

INPUT_PDF = Path("Current-Test/MBP 1 MK 2026.pdf")
# INPUT_PDF = Path("MBP 1 GG 2026_enhanced_reduced.pdf")
OUTPUT_XLSX = Path("extracted_tables.xlsx")
DEBUG_DIR = Path(__file__).resolve().parent / "table_debug"

DETECTION_MODEL_NAME = "microsoft/table-transformer-detection"
STRUCTURE_MODEL_NAME = (
    "microsoft/table-transformer-structure-recognition-v1.1-all"
)

PDF_DPI = 250
DETECTION_THRESHOLD = 0.75
STRUCTURE_THRESHOLD = 0.60

TABLE_PADDING = 12
CELL_INSET = 2

# Uncomment and update this on Windows only when tesseract.exe is not in PATH.
# This is unnecessary after replacing default_ocr() with your existing OCR.
#
pytesseract.pytesseract.tesseract_cmd = (
    r"C:\Users\Divyesh Parmar\Downloads\Tesserect OCR\tesseract.exe"
)


@dataclass(frozen=True)
class Detection:
    label: str
    score: float
    box: tuple[float, float, float, float]


# =============================================================================
# GENERAL GEOMETRY
# =============================================================================

def safe_sheet_name(name: str) -> str:
    """Return a legal Excel worksheet name."""
    name = re.sub(r"[:\\/?*\[\]]", "_", name)
    return name[:31] or "Table"


def clamp_box(
    box: tuple[float, float, float, float],
    width: int,
    height: int,
    padding: int = 0,
     ) -> tuple[int, int, int, int]:
    x1, y1, x2, y2 = box

    return (
        max(0, int(round(x1)) - padding),
        max(0, int(round(y1)) - padding),
        min(width, int(round(x2)) + padding),
        min(height, int(round(y2)) + padding),
    )


def box_area(box: tuple[float, float, float, float]) -> float:
    x1, y1, x2, y2 = box
    return max(0.0, x2 - x1) * max(0.0, y2 - y1)


def intersection_area(
    box_a: tuple[float, float, float, float],
    box_b: tuple[float, float, float, float],
    ) -> float:
    ax1, ay1, ax2, ay2 = box_a
    bx1, by1, bx2, by2 = box_b

    x1 = max(ax1, bx1)
    y1 = max(ay1, by1)
    x2 = min(ax2, bx2)
    y2 = min(ay2, by2)

    return max(0.0, x2 - x1) * max(0.0, y2 - y1)


def box_iou(
    box_a: tuple[float, float, float, float],
    box_b: tuple[float, float, float, float],
    ) -> float:
    intersection = intersection_area(box_a, box_b)
    union = box_area(box_a) + box_area(box_b) - intersection
    return intersection / union if union > 0 else 0.0


def remove_duplicate_detections(
    detections: list[Detection],
    iou_threshold: float,
    ) -> list[Detection]:
    """
    Remove near-identical model boxes while retaining the highest-confidence box.
    """
    kept: list[Detection] = []

    for item in sorted(detections, key=lambda detection: detection.score, reverse=True):
        duplicate = any(
            item.label == old.label
            and box_iou(item.box, old.box) >= iou_threshold
            for old in kept
        )
        if not duplicate:
            kept.append(item)

    return kept


# =============================================================================
# PDF RENDERING
# =============================================================================

def render_pdf_pages(pdf_path: Path, dpi: int) -> list[Image.Image]:
    """
    Render all PDF pages to RGB PIL images.

    A higher DPI improves small-text OCR and structure recognition, but increases
    processing time and memory usage.
    """
    images: list[Image.Image] = []
    zoom = dpi / 72.0
    matrix = fitz.Matrix(zoom, zoom)

    with fitz.open(pdf_path) as document:
        for page in document:
            pixmap = page.get_pixmap(matrix=matrix, alpha=False)
            image = Image.frombytes(
                "RGB",
                (pixmap.width, pixmap.height),
                pixmap.samples,
            )
            images.append(image)

    return images


# =============================================================================
# TABLE TRANSFORMER
# =============================================================================

class TableTransformerEngine:
    def __init__(self) -> None:
        self.device = torch.device(
            "cuda" if torch.cuda.is_available() else "cpu"
        )

        print(f"Using device: {self.device}")

        self.detection_processor = AutoImageProcessor.from_pretrained(
            DETECTION_MODEL_NAME
        )
        self.detection_model = (
            TableTransformerForObjectDetection.from_pretrained(
                DETECTION_MODEL_NAME
            )
            .to(self.device)
            .eval()
        )

        self.structure_processor = AutoImageProcessor.from_pretrained(
            STRUCTURE_MODEL_NAME
        )
        # transformers>=4.4x expects both shortest_edge and longest_edge.
        # v1.1 structure configs only ship longest_edge and crash on resize.
        self._normalize_processor_size(self.detection_processor)
        self._normalize_processor_size(self.structure_processor)

        self.structure_model = (
            TableTransformerForObjectDetection.from_pretrained(
                STRUCTURE_MODEL_NAME
            )
            .to(self.device)
            .eval()
        )

    @staticmethod
    def _normalize_processor_size(processor) -> None:
        size = getattr(processor, "size", None)
        if not isinstance(size, dict):
            return

        if "shortest_edge" in size and "longest_edge" in size:
            return
        if "height" in size and "width" in size:
            return

        longest = int(size.get("longest_edge") or size.get("max_size") or 800)
        shortest = int(size.get("shortest_edge") or min(800, longest))
        processor.size = {
            "shortest_edge": shortest,
            "longest_edge": longest,
        }

    @staticmethod
    def _decode_results(
        result: dict,
        id2label: dict,
    ) -> list[Detection]:
        decoded: list[Detection] = []

        for score, label_tensor, box_tensor in zip(
            result["scores"],
            result["labels"],
            result["boxes"],
        ):
            label_id = int(label_tensor.item())
            label = id2label.get(
                label_id,
                id2label.get(str(label_id), str(label_id)),
            )

            decoded.append(
                Detection(
                    label=str(label).lower(),
                    score=float(score.item()),
                    box=tuple(
                        float(coordinate)
                        for coordinate in box_tensor.tolist()
                    ),
                )
            )

        return decoded

    @staticmethod
    def _move_inputs_to_device(inputs: dict, device: torch.device) -> dict:
        return {
            key: value.to(device)
            for key, value in inputs.items()
        }

    def detect_tables(self, page_image: Image.Image) -> list[Detection]:
        inputs = self.detection_processor(
            images=page_image,
            return_tensors="pt",
        )
        inputs = self._move_inputs_to_device(inputs, self.device)

        with torch.inference_mode():
            outputs = self.detection_model(**inputs)

        target_sizes = torch.tensor(
            [[page_image.height, page_image.width]],
            device=self.device,
        )

        result = self.detection_processor.post_process_object_detection(
            outputs,
            threshold=DETECTION_THRESHOLD,
            target_sizes=target_sizes,
        )[0]

        detections = self._decode_results(
            result,
            self.detection_model.config.id2label,
        )

        detections = [
            item
            for item in detections
            if item.label in {"table", "table rotated"}
        ]

        return remove_duplicate_detections(
            detections,
            iou_threshold=0.70,
        )

    def recognize_structure(
        self,
        table_image: Image.Image,
    ) -> list[Detection]:
        inputs = self.structure_processor(
            images=table_image,
            return_tensors="pt",
        )
        inputs = self._move_inputs_to_device(inputs, self.device)

        with torch.inference_mode():
            outputs = self.structure_model(**inputs)

        target_sizes = torch.tensor(
            [[table_image.height, table_image.width]],
            device=self.device,
        )

        result = self.structure_processor.post_process_object_detection(
            outputs,
            threshold=STRUCTURE_THRESHOLD,
            target_sizes=target_sizes,
        )[0]

        return self._decode_results(
            result,
            self.structure_model.config.id2label,
        )


# =============================================================================
# OCR ADAPTER
# =============================================================================

def default_ocr(cell_image: Image.Image) -> str:
    """
    Standalone OCR implementation using Tesseract.

    Replace only this function with your existing OCR implementation.
    The replacement must:
      1. Accept a PIL.Image.Image.
      2. Return the recognized text as a string.
    """
    if cell_image.width < 2 or cell_image.height < 2:
        return ""

    enlarged = cell_image.resize(
        (cell_image.width * 2, cell_image.height * 2),
        Image.Resampling.LANCZOS,
    )

    text = pytesseract.image_to_string(
        enlarged,
        config="--oem 3 --psm 6",
    )

    # Preserve lines, but normalize unnecessary spaces.
    normalized_lines = [
        " ".join(line.split())
        for line in text.splitlines()
        if line.strip()
    ]

    return "\n".join(normalized_lines).strip()


# Example adapter for your existing OCR:
#
# def my_existing_ocr(cell_image: Image.Image) -> str:
#     result = your_ocr_model.recognize(cell_image)
#     return result["text"].strip()
#
# Then change the last function call to:
#
# extract_tables_from_pdf(
#     INPUT_PDF,
#     OUTPUT_XLSX,
#     ocr_function=my_existing_ocr,
# )


# =============================================================================
# TABLE STRUCTURE TO GRID
# =============================================================================

def prepare_rows_columns_and_spans(
    structure: list[Detection],
    ) -> tuple[list[Detection], list[Detection], list[Detection]]:
    rows = [
        item
        for item in structure
        if item.label == "table row"
    ]
    columns = [
        item
        for item in structure
        if item.label == "table column"
    ]
    spans = [
        item
        for item in structure
        if item.label == "table spanning cell"
    ]

    rows = remove_duplicate_detections(rows, iou_threshold=0.85)
    columns = remove_duplicate_detections(columns, iou_threshold=0.85)
    spans = remove_duplicate_detections(spans, iou_threshold=0.85)

    rows.sort(key=lambda item: (item.box[1] + item.box[3]) / 2)
    columns.sort(key=lambda item: (item.box[0] + item.box[2]) / 2)

    return rows, columns, spans


def grid_cell_box(
    row: Detection,
    column: Detection,
    image_width: int,
    image_height: int,
    ) -> tuple[int, int, int, int]:
    """
    Build a cell using the intersection of one predicted row and one column.
    """
    x1 = max(0, int(round(column.box[0])))
    y1 = max(0, int(round(row.box[1])))
    x2 = min(image_width, int(round(column.box[2])))
    y2 = min(image_height, int(round(row.box[3])))

    if x2 <= x1 or y2 <= y1:
        return (0, 0, 0, 0)

    return (x1, y1, x2, y2)


def find_span_range(
    span: Detection,
    rows: list[Detection],
    columns: list[Detection],
    table_width: int,
    table_height: int,
    minimum_cell_overlap: float = 0.35,
    ) -> tuple[int, int, int, int] | None:
    """
    Convert a predicted spanning-cell box into zero-based grid coordinates.

    Returns:
        row_start, row_end, column_start, column_end
    """
    matched_rows: set[int] = set()
    matched_columns: set[int] = set()

    for row_index, row in enumerate(rows):
        for column_index, column in enumerate(columns):
            cell_box = grid_cell_box(
                row,
                column,
                table_width,
                table_height,
            )
            area = box_area(cell_box)

            if area <= 0:
                continue

            overlap_ratio = intersection_area(span.box, cell_box) / area

            if overlap_ratio >= minimum_cell_overlap:
                matched_rows.add(row_index)
                matched_columns.add(column_index)

    if not matched_rows or not matched_columns:
        return None

    row_start = min(matched_rows)
    row_end = max(matched_rows)
    column_start = min(matched_columns)
    column_end = max(matched_columns)

    if row_start == row_end and column_start == column_end:
        return None

    return row_start, row_end, column_start, column_end


def ranges_overlap(
    first: tuple[int, int, int, int],
    second: tuple[int, int, int, int],
    ) -> bool:
    first_row_start, first_row_end, first_col_start, first_col_end = first
    second_row_start, second_row_end, second_col_start, second_col_end = second

    rows_overlap = not (
        first_row_end < second_row_start
        or second_row_end < first_row_start
    )
    columns_overlap = not (
        first_col_end < second_col_start
        or second_col_end < first_col_start
    )

    return rows_overlap and columns_overlap


def extract_grid(
    table_image: Image.Image,
    rows: list[Detection],
    columns: list[Detection],
    spans: list[Detection],
    ocr_function: Callable[[Image.Image], str],
    ) -> tuple[
    list[list[str]],
    list[tuple[int, int, int, int]],
    ]:
    """
    OCR every ordinary cell, then replace predicted spanning ranges with one
    OCR result for the complete merged region.
    """
    data = [
        ["" for _ in columns]
        for _ in rows
    ]

    for row_index, row in enumerate(rows):
        for column_index, column in enumerate(columns):
            x1, y1, x2, y2 = grid_cell_box(
                row,
                column,
                table_image.width,
                table_image.height,
            )

            if x2 <= x1 or y2 <= y1:
                continue

            x1 = min(x2, x1 + CELL_INSET)
            y1 = min(y2, y1 + CELL_INSET)
            x2 = max(x1, x2 - CELL_INSET)
            y2 = max(y1, y2 - CELL_INSET)

            cell_image = table_image.crop((x1, y1, x2, y2))
            data[row_index][column_index] = ocr_function(cell_image)

    merge_ranges: list[tuple[int, int, int, int]] = []

    for span in sorted(
        spans,
        key=lambda item: item.score,
        reverse=True,
    ):
        merge_range = find_span_range(
            span,
            rows,
            columns,
            table_image.width,
            table_image.height,
        )

        if merge_range is None:
            continue

        if any(
            ranges_overlap(merge_range, existing)
            for existing in merge_ranges
        ):
            continue

        row_start, row_end, column_start, column_end = merge_range

        x1 = min(
            columns[index].box[0]
            for index in range(column_start, column_end + 1)
        )
        x2 = max(
            columns[index].box[2]
            for index in range(column_start, column_end + 1)
        )
        y1 = min(
            rows[index].box[1]
            for index in range(row_start, row_end + 1)
        )
        y2 = max(
            rows[index].box[3]
            for index in range(row_start, row_end + 1)
        )

        merged_box = clamp_box(
            (x1, y1, x2, y2),
            table_image.width,
            table_image.height,
        )
        merged_image = table_image.crop(merged_box)
        merged_text = ocr_function(merged_image)

        data[row_start][column_start] = merged_text

        for row_index in range(row_start, row_end + 1):
            for column_index in range(column_start, column_end + 1):
                if (
                    row_index != row_start
                    or column_index != column_start
                ):
                    data[row_index][column_index] = ""

        merge_ranges.append(merge_range)

    return data, merge_ranges


# =============================================================================
# DEBUG IMAGES
# =============================================================================

def save_structure_debug_image(
    table_image: Image.Image,
    rows: list[Detection],
    columns: list[Detection],
    spans: list[Detection],
    output_path: Path,
    ) -> None:
    """
    Red = rows, blue = columns, green = spanning cells.
    """
    debug_image = table_image.copy()
    draw = ImageDraw.Draw(debug_image)

    for row in rows:
        draw.rectangle(row.box, outline="red", width=2)

    for column in columns:
        draw.rectangle(column.box, outline="blue", width=2)

    for span in spans:
        draw.rectangle(span.box, outline="green", width=3)

    debug_image.save(output_path)


# =============================================================================
# EXCEL OUTPUT
# =============================================================================

def write_table_sheet(
    workbook: Workbook,
    sheet_name: str,
    data: list[list[str]],
    merge_ranges: list[tuple[int, int, int, int]],
     ) -> None:
    worksheet = workbook.create_sheet(
        title=safe_sheet_name(sheet_name)
    )
    worksheet.freeze_panes = "A2"

    for row_number, values in enumerate(data, start=1):
        for column_number, value in enumerate(values, start=1):
            cell = worksheet.cell(
                row=row_number,
                column=column_number,
                value=value,
            )
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )

    # Treat the first detected row as a probable header.
    if data:
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="D9EAF7",
            )

    # Determine widths before applying Excel merges.
    for column_number in range(1, len(data[0]) + 1 if data else 1):
        maximum_length = 0

        for row_number in range(1, len(data) + 1):
            value = worksheet.cell(
                row=row_number,
                column=column_number,
            ).value

            if value is None:
                continue

            maximum_length = max(
                maximum_length,
                max(
                    (
                        len(line)
                        for line in str(value).splitlines()
                    ),
                    default=0,
                ),
            )

        worksheet.column_dimensions[
            get_column_letter(column_number)
        ].width = min(max(maximum_length + 2, 10), 45)

    for row_start, row_end, column_start, column_end in merge_ranges:
        worksheet.merge_cells(
            start_row=row_start + 1,
            end_row=row_end + 1,
            start_column=column_start + 1,
            end_column=column_end + 1,
        )


# =============================================================================
# MAIN PIPELINE
# =============================================================================

def extract_tables_from_pdf(
    input_pdf: Path,
    output_xlsx: Path,
    ocr_function: Callable[[Image.Image], str] = default_ocr,
    ) -> None:
    if not input_pdf.exists():
        raise FileNotFoundError(
            f"Input PDF was not found: {input_pdf.resolve()}"
        )

    DEBUG_DIR.mkdir(parents=True, exist_ok=True)

    engine = TableTransformerEngine()
    pages = render_pdf_pages(input_pdf, PDF_DPI)

    workbook = Workbook()
    workbook.remove(workbook.active)

    extracted_table_count = 0

    for page_number, page_image in enumerate(pages, start=1):
        tables = engine.detect_tables(page_image)

        print(
            f"Page {page_number}: "
            f"{len(tables)} table region(s) detected."
        )

        for table_number, table in enumerate(tables, start=1):
            crop_box = clamp_box(
                table.box,
                page_image.width,
                page_image.height,
                padding=TABLE_PADDING,
            )
            table_image = page_image.crop(crop_box)

            if table.label == "table rotated":
                table_image = table_image.rotate(
                    90,
                    expand=True,
                )

            table_crop_path = (
                DEBUG_DIR
                / f"page_{page_number}_table_{table_number}.png"
            )
            DEBUG_DIR.mkdir(parents=True, exist_ok=True)
            table_image.save(table_crop_path)

            structure = engine.recognize_structure(table_image)
            rows, columns, spans = prepare_rows_columns_and_spans(
                structure
            )

            structure_debug_path = DEBUG_DIR / (
                f"page_{page_number}_table_{table_number}"
                "_structure.png"
            )
            DEBUG_DIR.mkdir(parents=True, exist_ok=True)
            save_structure_debug_image(
                table_image,
                rows,
                columns,
                spans,
                structure_debug_path,
            )

            if not rows or not columns:
                print(
                    f"  Skipped table {table_number}: "
                    "no usable row/column structure."
                )
                continue

            data, merge_ranges = extract_grid(
                table_image,
                rows,
                columns,
                spans,
                ocr_function,
            )

            write_table_sheet(
                workbook,
                sheet_name=f"P{page_number}_Table{table_number}",
                data=data,
                merge_ranges=merge_ranges,
            )

            extracted_table_count += 1
            print(
                f"  Extracted table {table_number}: "
                f"{len(rows)} row(s), {len(columns)} column(s)."
            )

    if extracted_table_count == 0:
        worksheet = workbook.create_sheet("No tables found")
        worksheet["A1"] = (
            "No table with usable row and column structure was detected."
        )

    workbook.save(output_xlsx)

    print()
    print(f"Tables extracted: {extracted_table_count}")
    print(f"Excel file: {output_xlsx.resolve()}")
    print(f"Debug images: {DEBUG_DIR.resolve()}")


if __name__ == "__main__":
    extract_tables_from_pdf(
        input_pdf=INPUT_PDF,
        output_xlsx=OUTPUT_XLSX,
        ocr_function=default_ocr,
    )
