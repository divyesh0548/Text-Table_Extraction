import os
import warnings
import cv2
import fitz
import shutil
import numpy as np
from pdf2image import convert_from_path
from PIL import Image
import img2pdf
import unstructured_pytesseract
import pandas as pd
from unstructured.partition.pdf import partition_pdf
from openpyxl import Workbook, load_workbook

# Configure warnings and tesseract path
warnings.filterwarnings('ignore')
unstructured_pytesseract.pytesseract.tesseract_cmd = r'C:\Users\Divyesh Parmar\Downloads\Tesserect OCR\tesseract.exe'

# ========== CONFIGURATION ==========
PDF_PATH = "CH-012026-0062-INVOICE.pdf"  # relative to current working directory
PAGES_TO_PROCESS = "1"           # None for all pages, or e.g. "1,3-5"
ROTATE_PDF = False
ROTATION_PAGES = None            # None to use processed pages, or e.g. "1,2"
ROTATION_ANGLE = 90                # 90, 180, 270, or -90
ENHANCE_PDF = False
KEEP_FINAL_PDF = False           # Keep the final processed PDF after extraction
USE_INTERACTIVE_INPUT = False    # Set True to prompt for inputs instead of using config above


def cleanup_temp_files():
    """Remove temporary files and directories generated during processing"""
    temp_patterns = ['temp_pdf_images', 'figures', 'figure-']

    # Remove temp directories
    for pattern in temp_patterns:
        if os.path.exists(pattern):
            shutil.rmtree(pattern, ignore_errors=True)
            print(f"[CLEANUP] Removed {pattern} directory")

    # Remove temp files
    current_dir = os.getcwd()
    for filename in os.listdir(current_dir):
        if ('temp_page_' in filename and filename.endswith('.jpg')) or \
           ('temp_' in filename and filename.endswith('.png')) or \
           filename.endswith('_temp.pdf') or \
           (filename.startswith('output_table_') and filename.endswith('.xlsx')) or \
           (filename.startswith('figure-') and filename.endswith(('.png', '.jpg', '.jpeg'))):
            try:
                os.remove(filename)
                print(f"[CLEANUP] Removed {filename}")
            except Exception as e:
                print(f"[WARN] Could not delete {filename}: {e}")


def clean_excel_data(excel_path):
    """Advanced data cleaning for Excel file"""
    print(f"\n[PROCESS] Starting advanced data cleaning on {excel_path}...")

    try:
        # Load the workbook
        wb = load_workbook(excel_path)

        processed_sheets = 0
        total_cells_processed = 0
        nan_removed = 0
        duplicates_removed = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"[PROCESS] Processing sheet: {sheet_name}")

            # Get sheet dimensions
            max_row = ws.max_row
            max_col = ws.max_column

            if max_row == 1 and max_col == 1:  # Empty sheet
                continue

            # Step 1: Remove all 'nan', 'NaN', 'None', empty strings
            cells_cleaned_nan = 0
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None:
                        cell_value = str(cell.value).strip().lower()
                        if cell_value in ['nan', 'none', '', 'null']:
                            cell.value = None
                            cells_cleaned_nan += 1
                            nan_removed += 1

            print(f"[PROCESS] Removed {cells_cleaned_nan} NaN/empty values from {sheet_name}")

            # Step 2: Remove duplicates in rows (left to right priority)
            duplicates_in_rows = 0
            for row in range(1, max_row + 1):
                seen_values = {}  # {value: first_column_position}

                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None and str(cell.value).strip():
                        cell_value = str(cell.value).strip()

                        if cell_value in seen_values:
                            # Found duplicate, check if there's gap (empty cell) between
                            first_col = seen_values[cell_value]
                            has_gap = any(
                                ws.cell(row=row, column=c).value is None or 
                                str(ws.cell(row=row, column=c).value).strip() == ""
                                for c in range(first_col + 1, col)
                            )

                            # Remove duplicate (keep left, remove right)
                            cell.value = None
                            duplicates_in_rows += 1
                            duplicates_removed += 1
                        else:
                            seen_values[cell_value] = col

            print(f"[PROCESS] Removed {duplicates_in_rows} row duplicates from {sheet_name}")

            # Step 3: Remove duplicates in columns (top to bottom priority)
            duplicates_in_cols = 0
            for col in range(1, max_col + 1):
                seen_values = {}  # {value: first_row_position}

                for row in range(1, max_row + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None and str(cell.value).strip():
                        cell_value = str(cell.value).strip()

                        if cell_value in seen_values:
                            # Found duplicate, check if there's gap (empty cell) between
                            first_row = seen_values[cell_value]
                            has_gap = any(
                                ws.cell(row=r, column=col).value is None or 
                                str(ws.cell(row=r, column=col).value).strip() == ""
                                for r in range(first_row + 1, row)
                            )

                            # Remove duplicate (keep top, remove bottom)
                            cell.value = None
                            duplicates_in_cols += 1
                            duplicates_removed += 1
                        else:
                            seen_values[cell_value] = row

            print(f"[PROCESS] Removed {duplicates_in_cols} column duplicates from {sheet_name}")

            # Count total cells processed
            total_cells_processed += max_row * max_col
            processed_sheets += 1

        # Save the cleaned workbook
        cleaned_path = excel_path.replace('.xlsx', '_cleaned.xlsx')
        wb.save(cleaned_path)

        print(f"[PROCESS] Data cleaning completed!")
        print(f"[PROCESS] Sheets processed: {processed_sheets}")
        print(f"[PROCESS] Total cells processed: {total_cells_processed}")
        print(f"[PROCESS] NaN values removed: {nan_removed}")
        print(f"[PROCESS] Duplicate values removed: {duplicates_removed}")
        print(f"[PROCESS] Cleaned file saved as: {cleaned_path}")

        # Remove original file and rename cleaned file
        try:
            os.remove(excel_path)
            os.rename(cleaned_path, excel_path)
            print(f"[PROCESS] Original file replaced with cleaned version")
        except Exception as e:
            print(f"[WARN] Could not replace original file: {e}")
            return cleaned_path

        return excel_path

    except Exception as e:
        print(f"[ERROR] Data cleaning failed: {e}")
        return excel_path


def get_pdf_page_count(pdf_path):
    """Get total number of pages in PDF"""
    try:
        doc = fitz.open(pdf_path)
        total_pages = doc.page_count
        doc.close()
        return total_pages
    except Exception as e:
        print(f"Error getting page count: {e}")
        return 0


def parse_page_input(page_input, total_pages):
    """Parse page input like '1,3-5,7' into list of page numbers (0-indexed)"""
    pages = set()
    for part in page_input.split(','):
        part = part.strip()
        if '-' in part:
            start, end = part.split('-', 1)
            try:
                s, e = int(start)-1, int(end)-1
                pages.update(range(max(s, 0), min(e, total_pages-1)+1))
            except ValueError:
                continue
        else:
            try:
                p = int(part)-1
                if 0 <= p < total_pages:
                    pages.add(p)
            except ValueError:
                continue
    return sorted(pages) if pages else None


def get_user_input():
    """Get processing parameters from config variables or interactive prompts."""
    print("=== PDF Table Extraction Automation with Unstructured OCR ===\n")

    if USE_INTERACTIVE_INPUT:
        while True:
            pdf_path = input("Enter the path to your PDF file: ").strip().strip('"\'')
            if os.path.exists(pdf_path) and pdf_path.lower().endswith('.pdf'):
                break
            print("PDF file not found. Please enter a valid path.")
    else:
        pdf_path = PDF_PATH.strip().strip('"\'')
        if not os.path.exists(pdf_path) or not pdf_path.lower().endswith('.pdf'):
            raise FileNotFoundError(f"PDF file not found: {pdf_path}")

    total_pages = get_pdf_page_count(pdf_path)
    print(f"\nDetected {total_pages} pages in the PDF")

    if USE_INTERACTIVE_INPUT:
        page_input = input(f"Enter pages to process (1-{total_pages}, e.g., '1,3-5') or press Enter for all: ").strip()
        pages_to_process = parse_page_input(page_input, total_pages) if page_input else None

        rotate_pdf = input("\nDo you want to rotate PDF pages? (y/n): ").strip().lower() == 'y'
        rotation_info = {}

        if rotate_pdf:
            rotation_pages_input = input("Enter pages to rotate (same format as above) or press Enter for all processed pages: ").strip()
            rotation_pages = parse_page_input(rotation_pages_input, total_pages) if rotation_pages_input else pages_to_process

            if rotation_pages:
                while True:
                    try:
                        rotation_angle = int(input("Enter rotation angle (90, 180, 270, -90): "))
                        if rotation_angle in [90, 180, 270, -90]:
                            break
                        print("Please enter a valid rotation angle (90, 180, 270, -90)")
                    except ValueError:
                        print("Please enter a valid number")

                rotation_info = {'pages': rotation_pages, 'angle': rotation_angle}

        enhance_pdf = input("\nDo you want to enhance the PDF? (y/n): ").strip().lower() == 'y'
    else:
        pages_to_process = parse_page_input(PAGES_TO_PROCESS, total_pages) if PAGES_TO_PROCESS else None
        rotation_info = {}

        if ROTATE_PDF:
            rotation_pages = parse_page_input(ROTATION_PAGES, total_pages) if ROTATION_PAGES else pages_to_process
            if rotation_pages:
                if ROTATION_ANGLE not in [90, 180, 270, -90]:
                    raise ValueError(f"Invalid ROTATION_ANGLE: {ROTATION_ANGLE}. Use 90, 180, 270, or -90.")
                rotation_info = {'pages': rotation_pages, 'angle': ROTATION_ANGLE}

        enhance_pdf = ENHANCE_PDF

    return pdf_path, pages_to_process, rotation_info, enhance_pdf


def rotate_pdf_pages(pdf_path, rotation_info):
    """Rotate specific pages of PDF"""
    if not rotation_info:
        return pdf_path

    print(f"\n[ROTATE] Rotating pages {[p+1 for p in rotation_info['pages']]} by {rotation_info['angle']} degrees...")

    try:
        doc = fitz.open(pdf_path)

        for page_num in rotation_info['pages']:
            if 0 <= page_num < doc.page_count:
                page = doc[page_num]
                page.set_rotation(rotation_info['angle'])
                print(f"[ROTATE] Rotated page {page_num + 1}")

        base_name = os.path.splitext(pdf_path)[0]
        rotated_path = f"{base_name}_rotated.pdf"
        doc.save(rotated_path)
        doc.close()

        print(f"[ROTATE] Rotated PDF saved as: {rotated_path}")
        return rotated_path

    except Exception as e:
        print(f"[ERROR] Failed to rotate PDF: {e}")
        return pdf_path


def enhance_pdf_tables(pdf_path, dpi=300, line_strength=2, show_progress=True):
    """Enhanced PDF table enhancement function"""
    if show_progress:
        print(f"\n[ENHANCE] Starting PDF enhancement...")

    # Temporary folder for page images
    temp_dir = "temp_pdf_images"
    os.makedirs(temp_dir, exist_ok=True)

    try:
        # Convert PDF pages to images
        pages = convert_from_path(pdf_path, dpi=dpi)
        enhanced_images = []

        for i, page in enumerate(pages):
            if show_progress:
                print(f"[ENHANCE] Processing page {i+1}/{len(pages)}...")

            # Convert PIL image to OpenCV format
            img = np.array(page)
            img = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)

            # --- Step 1: Grayscale ---
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

            # --- Step 2: Improve contrast ---
            gray = cv2.convertScaleAbs(gray, alpha=1.5, beta=0)

            # --- Step 3: Reduce noise ---
            gray = cv2.GaussianBlur(gray, (3, 3), 0)

            # --- Step 4: Adaptive threshold ---
            thresh = cv2.adaptiveThreshold(
                gray, 255,
                cv2.ADAPTIVE_THRESH_MEAN_C,
                cv2.THRESH_BINARY_INV,
                15, 10
            )

            # --- Step 5: Detect horizontal & vertical lines ---
            horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (40, 1))
            vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 40))

            detect_horizontal = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, horizontal_kernel, iterations=2)
            detect_vertical = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, vertical_kernel, iterations=2)

            # --- Step 6: Combine and strengthen lines ---
            table_mask = cv2.add(detect_horizontal, detect_vertical)
            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (line_strength, line_strength))
            table_mask = cv2.dilate(table_mask, kernel, iterations=2)

            # --- Step 7: Smooth edges and repair small gaps ---
            table_mask = cv2.morphologyEx(table_mask, cv2.MORPH_CLOSE, kernel, iterations=1)

            # --- Step 8: Invert and merge with original ---
            inverted_mask = cv2.bitwise_not(table_mask)
            enhanced_gray = cv2.bitwise_and(gray, gray, mask=inverted_mask)
            enhanced_bgr = cv2.cvtColor(enhanced_gray, cv2.COLOR_GRAY2BGR)

            # Save temporary enhanced image
            enhanced_path = os.path.join(temp_dir, f"enhanced_page_{i+1}.png")
            cv2.imwrite(enhanced_path, enhanced_bgr)
            enhanced_images.append(Image.open(enhanced_path).convert("RGB"))

        # --- Step 9: Save final enhanced PDF ---
        base_name = os.path.splitext(os.path.basename(pdf_path))[0]
        pdf_output_path = os.path.join(os.getcwd(), f"{base_name}_enhanced.pdf")
        enhanced_images[0].save(pdf_output_path, save_all=True, append_images=enhanced_images[1:])

        if show_progress:
            print(f"[ENHANCE] Enhanced PDF saved as: {pdf_output_path}")

        return pdf_output_path

    except Exception as e:
        print(f"[ERROR] Enhancement failed: {e}")
        return pdf_path


def reduce_pdf_resolution(input_pdf_path, target_dpi=72):
    """Reduce PDF resolution"""
    print(f"\n[REDUCE] Reducing PDF resolution to {target_dpi} DPI...")

    try:
        # Convert PDF to images with original resolution
        images = convert_from_path(input_pdf_path)
        reduced_images = []

        for i, img in enumerate(images):
            # Calculate new size based on target DPI and original DPI
            orig_dpi = img.info.get('dpi', (300,300))[0]
            scale_factor = target_dpi / orig_dpi
            new_size = (int(img.width * scale_factor), int(img.height * scale_factor))

            # Resize image to new target resolution
            reduced_img = img.resize(new_size, Image.LANCZOS)

            # Save reduced image temporarily in JPEG format to reduce size
            temp_img_path = f"temp_page_{i}.jpg"
            reduced_img.save(temp_img_path, "JPEG", quality=85)
            reduced_images.append(temp_img_path)

        # Convert list of JPEG images back to PDF
        base_name = os.path.splitext(os.path.basename(input_pdf_path))[0]
        output_pdf_path = f"{base_name}_reduced.pdf"

        with open(output_pdf_path, "wb") as f:
            f.write(img2pdf.convert(reduced_images))

        # Clean up temporary images
        for img_path in reduced_images:
            os.remove(img_path)

        print(f"[REDUCE] Reduced resolution PDF saved as: {output_pdf_path}")
        return output_pdf_path

    except Exception as e:
        print(f"[ERROR] Resolution reduction failed: {e}")
        return input_pdf_path


def extract_tables_with_unstructured(pdf_path, pages=None):
    """Extract tables using Unstructured OCR"""
    print(f"\n[EXTRACT] Starting table extraction with Unstructured OCR...")

    try:
        # Convert 0-indexed pages to 1-indexed for partition_pdf
        page_list = [p + 1 for p in pages] if pages else None

        # Partition the PDF with page specification
        elements = partition_pdf(
            filename=pdf_path,
            infer_table_structure=True,
            extract_images_in_pdf=True,
            languages=["eng"],
            extract_image_block_types=["Table"],
            strategy='hi_res',
            pages=page_list
        )

        # Extract tables from elements
        tables = [el for el in elements if el.category == "Table"]

        if not tables:
            print("[EXTRACT] No tables found in the PDF")
            return None, 0

        print(f"[EXTRACT] Found {len(tables)} table(s)")

        # Create a single Excel workbook with multiple sheets
        wb = Workbook()

        # Remove default sheet if we have tables
        if wb.active.title == 'Sheet':
            wb.remove(wb.active)

        table_count = 0

        for i, table in enumerate(tables):
            try:
                # Get HTML content of the table
                table_html = table.metadata.text_as_html

                # Parse HTML to DataFrame
                parsed_dfs = pd.read_html(table_html)

                if parsed_dfs:
                    df = parsed_dfs[0]
                    # Clean the data
                    df = df.replace(r'[\|\[\]]', '', regex=True)

                    # Create a new worksheet for each table
                    ws = wb.create_sheet(title=f"Table_{i+1}")

                    # Write column headers
                    for col_num, column_title in enumerate(df.columns, 1):
                        ws.cell(row=1, column=col_num, value=str(column_title))

                    # Write data rows
                    for row_num, row_data in enumerate(df.itertuples(index=False), start=2):
                        for col_num, value in enumerate(row_data, 1):
                            ws.cell(row=row_num, column=col_num, value=str(value) if value is not None else "")

                    table_count += 1
                    print(f"[EXTRACT] Added Table {i+1} with {len(df)} rows and {len(df.columns)} columns")

                else:
                    print(f"[EXTRACT] No data parsed from table {i+1}")

            except Exception as e:
                print(f"[ERROR] Error processing table {i+1}: {e}")
                continue

        if table_count > 0:
            # Save the workbook
            base_name = os.path.splitext(os.path.basename(pdf_path))[0]
            excel_filename = f"{base_name}_extracted_tables.xlsx"
            wb.save(excel_filename)
            print(f"[EXTRACT] Tables saved to: {excel_filename}")

            return excel_filename, table_count
        else:
            print("[EXTRACT] No tables were successfully processed")
            return None, 0

    except Exception as e:
        print(f"[ERROR] Table extraction failed: {e}")
        return None, 0


def main():
    """Main automation function"""
    try:
        # Get all user inputs
        pdf_path, pages_to_process, rotation_info, enhance_pdf = get_user_input()

        print(f"\n=== Processing Summary ===")
        print(f"PDF: {pdf_path}")
        print(f"Pages to process: {'All' if pages_to_process is None else [p+1 for p in pages_to_process]}")
        print(f"Rotation: {'Yes' if rotation_info else 'No'}")
        print(f"Enhancement: {'Yes' if enhance_pdf else 'No'}")

        current_pdf = pdf_path
        intermediate_files = []

        # Step 1: Rotate PDF if requested
        if rotation_info:
            current_pdf = rotate_pdf_pages(current_pdf, rotation_info)
            if current_pdf != pdf_path:
                intermediate_files.append(current_pdf)

        # Step 2: Enhance PDF if requested
        if enhance_pdf:
            current_pdf = enhance_pdf_tables(current_pdf)
            if not current_pdf.endswith('_enhanced.pdf'):
                print("[WARN] Enhancement may have failed, proceeding with current PDF")
            else:
                intermediate_files.append(current_pdf)

            # Step 3: Reduce resolution (only if enhanced)
            reduced_pdf = reduce_pdf_resolution(current_pdf)
            if reduced_pdf != current_pdf:
                current_pdf = reduced_pdf
                intermediate_files.append(current_pdf)

        # Step 4: Extract tables using Unstructured OCR
        excel_file, table_count = extract_tables_with_unstructured(current_pdf, pages_to_process)

        if excel_file and table_count > 0:
            # Step 5: Clean the Excel data (NEW FEATURE)
            cleaned_excel = clean_excel_data(excel_file)

            print(f"\n=== SUCCESS ===")
            print(f"Extracted and cleaned {table_count} tables to: {cleaned_excel}")

            if USE_INTERACTIVE_INPUT:
                keep_final_pdf = input(f"\nKeep the final processed PDF ({current_pdf})? (y/n): ").strip().lower() == 'y'
            else:
                keep_final_pdf = KEEP_FINAL_PDF

            # Cleanup intermediate files
            print(f"\n[CLEANUP] Cleaning up intermediate files...")
            for file_path in intermediate_files:
                if file_path != current_pdf or not keep_final_pdf:
                    try:
                        os.remove(file_path)
                        print(f"[CLEANUP] Removed {file_path}")
                    except Exception as e:
                        print(f"[WARN] Could not delete {file_path}: {e}")

            # Clean up temporary files
            cleanup_temp_files()

            if keep_final_pdf and current_pdf != pdf_path:
                print(f"[INFO] Final processed PDF kept as: {current_pdf}")

            print(f"\n=== FINAL OUTPUT ===")
            print(f"Clean Excel file with extracted tables: {cleaned_excel}")

        else:
            print(f"\n=== FAILED ===")
            print("No tables were extracted. Please check your PDF and try again.")

            # Still cleanup intermediate files
            for file_path in intermediate_files:
                try:
                    os.remove(file_path)
                except:
                    pass
            cleanup_temp_files()

    except KeyboardInterrupt:
        print(f"\n\n[INTERRUPTED] Process interrupted by user")
        cleanup_temp_files()
    except Exception as e:
        print(f"\n[ERROR] Unexpected error: {e}")
        cleanup_temp_files()


if __name__ == "__main__":
    main()
