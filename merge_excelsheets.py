import pandas as pd
from openpyxl import Workbook, load_workbook

# Load dataframes from the two Excel files
df1 = pd.read_excel('output_table_1.xlsx')
df2 = pd.read_excel('output_table_2.xlsx')

# Create a new workbook and select the active worksheet
wb = Workbook()
ws = wb.active


final_df = [df1, df2]
current_row = 1

for df in final_df:
    for col_num, column_title in enumerate(df.columns, 1):
        ws.cell(row=current_row, column=col_num, value=column_title)

    for row_num, row_data in enumerate(df.itertuples(index=False), start=2):
        current_row += 1
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=current_row, column=col_num, value=value)
    current_row += 2

wb.save('merged_output.xlsx')




# for col_num, column_title in enumerate(df.columns, 1):
#     ws.cell(row=1, column=col_num, value=column_title)

# for row_num, row_data in enumerate(df1.itertuples(index=False), start=2):
#     for col_num, value in enumerate(row_data, 1):
#         ws.cell(row=row_num, column=col_num, value=value)

# Calculate the start row for the second dataframe: 
# last row of df1 + 2 (one blank row)
# start_row = len(df1) + 3  # +1 for header, +1 blank row, +1 because rows start at 1

# # Write the second dataframe (df2) starting at start_row, including column headers
# for col_num, column_title in enumerate(df2.columns, 1):
#     ws.cell(row=start_row, column=col_num, value=column_title)

# for row_num, row_data in enumerate(df2.itertuples(index=False), start=start_row + 1):
#     for col_num, value in enumerate(row_data, 1):
#         ws.cell(row=row_num, column=col_num, value=value)

# # Save the new workbook to a file
# wb.save('merged_output.xlsx')

# print("Excel files merged successfully with table 2 below table 1.")
