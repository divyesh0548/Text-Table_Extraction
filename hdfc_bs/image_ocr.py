# UNSTRUCTURED LIBRARY: IMAGE-BASED PDF PROCESSING
# Solution for PDF text extraction restrictions using unstructured

import warnings
warnings.filterwarnings('ignore')
import unstructured_pytesseract

unstructured_pytesseract.pytesseract.tesseract_cmd = r'C:\Users\Divyesh Parmar\Downloads\Tesserect OCR\tesseract.exe'

import pandas as pd
import os
import tempfile
import shutil
from pathlib import Path

# Use unstructured for image processing instead of PDF
from unstructured.partition.image import partition_image
from openpyxl import Workbook, load_workbook

# PDF to Image conversion
import pdf2image

class UnstructuredImageProcessor:
    """
    Uses unstructured library to process PDF by converting to images first
    This bypasses PDF text extraction restrictions
    """
    
    def __init__(self):
        # Create temporary directory for images
        self.temp_dir = tempfile.mkdtemp(prefix="unstructured_images_")
        print(f"🗂️ Created temporary directory: {self.temp_dir}")
    
    def pdf_to_images(self, pdf_file, dpi=300):
        """
        Convert PDF to images using pdf2image
        """
        print(f"📄 Converting {pdf_file} to images...")
        
        try:
            # Convert PDF to images
            images = pdf2image.convert_from_path(
                pdf_file,
                dpi=dpi,
                output_folder=self.temp_dir,
                fmt='PNG'
            )
            
            image_paths = []
            for i, image in enumerate(images):
                image_path = os.path.join(self.temp_dir, f"page_{i+1:03d}.png")
                image.save(image_path, 'PNG')
                image_paths.append(image_path)
                print(f"   ✅ Page {i+1} → {os.path.basename(image_path)}")
            
            print(f"✅ Converted {len(image_paths)} pages to images")
            return image_paths
            
        except Exception as e:
            print(f"❌ PDF to image conversion failed: {e}")
            print("💡 Install poppler: conda install -c conda-forge poppler")
            return []
    
    def process_images_with_unstructured(self, image_paths):
        """
        Process each image using unstructured library
        """
        print(f"🔍 Processing {len(image_paths)} images with unstructured...")
        
        all_tables = []
        
        for i, image_path in enumerate(image_paths, 1):
            print(f"\n📄 Processing page {i}: {os.path.basename(image_path)}")
            
            try:
                # Use unstructured to partition the image
                elements = partition_image(
                    filename=image_path,
                    infer_table_structure=True,
                    extract_images_in_pdf=False,  # We're already working with images
                    languages=["eng"],
                    extract_image_block_types=["Table"],
                    strategy='hi_res'
                )
                
                # Extract tables from elements
                tables = [el for el in elements if el.category == "Table"]
                print(f"   Found {len(tables)} tables on page {i}")
                
                # Process each table
                for j, table in enumerate(tables):
                    try:
                        # Get HTML content of the table
                        table_html = table.metadata.text_as_html
                        
                        # Parse HTML to DataFrame
                        parsed_dfs = pd.read_html(table_html)
                        
                        if parsed_dfs:
                            df = parsed_dfs[0]
                            
                            # Clean the DataFrame
                            df = df.replace(r'[\|\[\]]', '', regex=True)
                            
                            print(f"   📊 Table {j+1}: {df.shape}")
                            
                            # Store table info
                            table_info = {
                                'page': i,
                                'table': j+1,
                                'dataframe': df,
                                'source': f'Page_{i}_Table_{j+1}'
                            }
                            
                            all_tables.append(table_info)
                        
                    except Exception as e:
                        print(f"   ❌ Error processing table {j+1}: {e}")
                        continue
            
            except Exception as e:
                print(f"   ❌ Error processing page {i}: {e}")
                continue
        
        print(f"\n✅ Total tables extracted: {len(all_tables)}")
        return all_tables
    
    def save_tables_to_excel(self, tables, output_filename="unstructured_extracted_tables.xlsx"):
        """
        Save tables to Excel with individual sheets and merged sheet
        """
        print(f"💾 Saving {len(tables)} tables to Excel...")
        
        if not tables:
            print("❌ No tables to save!")
            return None
        
        # Create individual Excel files first (like original code)
        excel_files = []
        
        for i, table_info in enumerate(tables):
            df = table_info['dataframe']
            
            if not df.empty:
                temp_filename = f"temp_table_{i+1}.xlsx"
                df.to_excel(temp_filename, index=False)
                excel_files.append(temp_filename)
                print(f"   Saved temp table {i+1}: {temp_filename}")
        
        # Read all Excel files and merge them (following original logic)
        dataframes = [pd.read_excel(f, engine='openpyxl') for f in excel_files]
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Merged_Tables"
        current_row = 1
        
        # Merge all dataframes into one sheet
        for i, df in enumerate(dataframes):
            # Add table header
            ws.cell(row=current_row, column=1, value=f"=== TABLE {i+1} FROM PAGE {tables[i]['page']} ===")
            current_row += 1
            
            # Add column headers
            for col_num, column_title in enumerate(df.columns, 1):
                ws.cell(row=current_row, column=col_num, value=column_title)
            current_row += 1
            
            # Add data rows
            for row_data in df.itertuples(index=False):
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=current_row, column=col_num, value=value)
                current_row += 1
            
            current_row += 2  # Add spacing between tables
        
        # Also create individual sheets for each table
        for i, table_info in enumerate(tables):
            df = table_info['dataframe']
            if not df.empty:
                sheet_name = f"Page{table_info['page']}_T{table_info['table']}"
                ws_individual = wb.create_sheet(title=sheet_name)
                
                # Write headers
                for col_num, column_title in enumerate(df.columns, 1):
                    ws_individual.cell(row=1, column=col_num, value=column_title)
                
                # Write data
                for row_num, row_data in enumerate(df.itertuples(index=False), start=2):
                    for col_num, value in enumerate(row_data, 1):
                        ws_individual.cell(row=row_num, column=col_num, value=value)
        
        wb.save(output_filename)
        print(f"✅ Saved final Excel: {output_filename}")
        
        # Clean up temporary Excel files (following original logic)
        for f in excel_files:
            try:
                os.remove(f)
                print(f"   Deleted temp file: {f}")
            except Exception as e:
                print(f"   Error deleting {f}: {e}")
        
        return output_filename
    
    def cleanup_temp_images(self):
        """
        Delete temporary image files
        """
        print(f"🧹 Cleaning up temporary images...")
        
        try:
            temp_files = os.listdir(self.temp_dir)
            print(f"   Deleting {len(temp_files)} temporary image files...")
            
            shutil.rmtree(self.temp_dir)
            print(f"✅ Deleted temporary directory: {os.path.basename(self.temp_dir)}")
        except Exception as e:
            print(f"⚠️ Could not delete temp directory: {e}")
    
    def process_pdf(self, pdf_file):
        """
        Main processing function that follows your original logic but with images
        """
        print("🚀 UNSTRUCTURED IMAGE-BASED PDF PROCESSING")
        print("="*50)
        print(f"Processing: {pdf_file}")
        print("This bypasses PDF text extraction restrictions by using images!")
        
        try:
            # Step 1: Convert PDF to images (bypasses restrictions)
            image_paths = self.pdf_to_images(pdf_file)
            
            if not image_paths:
                print("❌ Failed to convert PDF to images")
                return None
            
            # Step 2: Process images with unstructured (your original logic)
            tables = self.process_images_with_unstructured(image_paths)
            
            if not tables:
                print("❌ No tables extracted")
                self.cleanup_temp_images()
                return None
            
            # Step 3: Save to Excel (following your original structure)
            output_file = self.save_tables_to_excel(tables, 'unstructured_merged_output.xlsx')
            
            # Step 4: Clean up temporary images
            self.cleanup_temp_images()
            
            print(f"\n🎉 SUCCESS!")
            print(f"Processed {len(image_paths)} pages")
            print(f"Extracted {len(tables)} tables") 
            print(f"Output: {output_file}")
            print("All temporary files have been cleaned up.")
            
            return output_file
        
        except Exception as e:
            print(f"❌ Processing failed: {e}")
            self.cleanup_temp_images()
            return None

# MAIN FUNCTION - Direct replacement for your code
def process_restricted_pdf_with_unstructured(filename="bs1.pdf"):
    """
    Enhanced version of your original unstructured code
    Handles PDFs with text extraction restrictions
    """
    
    processor = UnstructuredImageProcessor()
    result = processor.process_pdf(filename)
    
    if result:
        print(f"\n✅ Your unstructured workflow completed successfully!")
        print(f"Output saved to: {result}")
    else:
        print(f"\n❌ Processing failed. Check requirements:")
        print("  - pdf2image: pip install pdf2image") 
        print("  - poppler: conda install -c conda-forge poppler")
        print("  - unstructured: pip install unstructured[local-inference]")
    
    return result

# EXACT REPLACEMENT FOR YOUR ORIGINAL CODE
def enhanced_unstructured_extraction(filename):
    
    # Process using image-based approach (bypasses restrictions)
    processor = UnstructuredImageProcessor()
    
    # Convert PDF to images first
    image_paths = processor.pdf_to_images(filename)
    
    if not image_paths:
        print("❌ Could not convert PDF to images")
        return
    
    # Process each image with unstructured (same logic as your original)
    dfs = []
    excel_files = []
    
    for i, image_path in enumerate(image_paths):
        print(f"\n🔍 Processing image {i+1}: {os.path.basename(image_path)}")
        
        # Call the partition_image function (instead of partition_pdf)
        # Returns a List[Element] present in the processed image
        elements = partition_image(filename=image_path,
                                 infer_table_structure=True,
                                 extract_images_in_pdf=False,  # Already working with images
                                 languages=["eng"],
                                 extract_image_block_types=["Table"],
                                 strategy='hi_res'
                                )
        
        tables = [el for el in elements if el.category == "Table"]
        
        for j, table in enumerate(tables):
            # Get HTML content of the table (same as your original)
            table_html = table.metadata.text_as_html
            # Parse HTML to DataFrame (same as your original) 
            parsed_dfs = pd.read_html(table_html)
            
            # Code for exporting into different excel and then combining them (same as your original)
            if parsed_dfs:
                df = parsed_dfs[0]
                
                df = df.replace(r'[\|\[\]]', '', regex=True)
                print(f"Table {len(excel_files)+1} Dataframe: \\n", df)
                
                output_filename = f"output_table_{len(excel_files)+1}.xlsx"
                df.to_excel(output_filename, index=False)
                excel_files.append(output_filename)
                print(f"Extracted table {len(excel_files)} saved to {output_filename}")
            
            else:
                print(f"No table parsed from element {j}")
    
    # Merging logic (exactly same as your original)
    dataframes = [pd.read_excel(f, engine='openpyxl') for f in excel_files]
    
    wb = Workbook()
    ws = wb.active
    current_row = 1
    
    for df in dataframes:
        for col_num, column_title in enumerate(df.columns, 1):
            ws.cell(row=current_row, column=col_num, value=column_title)
        
        for row_num, row_data in enumerate(df.itertuples(index=False), start=2):
            current_row += 1
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=current_row, column=col_num, value=value)
        current_row += 2
    
    wb.save('merged_output.xlsx')
    
    # Deleting temporary files (same as your original)
    for f in excel_files:
        try:
            os.remove(f)
            print(f"Deleted temp file: {f}")
        except Exception as e:
            print(f"Error deleting {f}: {e}")
    
    # Clean up image files
    processor.cleanup_temp_images()
    
    print("🎉 Enhanced unstructured extraction completed!")

# USAGE - Run this instead of your original code
if __name__ == "__main__":
    print("🔄 ENHANCED UNSTRUCTURED EXTRACTION")
    print("="*40)
    print("This uses your exact same unstructured logic but processes images instead of PDF")
    print("Completely bypasses PDF text extraction restrictions!")
    print()
    
    # enhanced_unstructured_extraction("bs1.pdf")
    
    process_restricted_pdf_with_unstructured("bs2.pdf")